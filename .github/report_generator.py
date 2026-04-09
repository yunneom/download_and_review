"""
통합 리포트 생성 모듈
- 여러 PID의 검증 결과를 하나의 Excel 파일로 통합
- PID별 요약 시트 + 개별 검증 결과 시트 구성
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from logger import logger


class IntegratedReportGenerator:
    """통합 리포트 생성기"""
    
    def __init__(self, download_date):
        """
        :param download_date: 다운로드 날짜 (폴더명, YYYY-MM-DD)
        """
        self.download_date = download_date
        self.validation_results = {}  # {pid: {date_range: {결과 정보}}}
        
    def add_validation_result(self, pid, date_range, result_file_path, status_summary):
        """
        검증 결과 추가
        :param pid: 차량 PID
        :param date_range: 날짜 범위 (예: "2026-01-09_to_2026-01-15")
        :param result_file_path: 검증 결과 파일 경로
        :param status_summary: 상태 요약 {'PASS': 10, 'FAIL': 2, 'WARNING': 1, 'N/A': 0}
        """
        if pid not in self.validation_results:
            self.validation_results[pid] = {}
        
        self.validation_results[pid][date_range] = {
            'file_path': result_file_path,
            'summary': status_summary,
            'overall_status': 'PASS' if status_summary['FAIL'] == 0 else 'FAIL'
        }
        
        logger.info(f"통합 리포트에 추가: PID {pid}, 기간 {date_range}")
    
    def generate_integrated_report(self, output_path):
        """
        통합 리포트 생성
        :param output_path: 출력 파일 경로
        """
        if not self.validation_results:
            logger.warning("통합 리포트에 추가할 데이터가 없습니다.")
            return
        
        wb = Workbook()
        
        # 1. 요약 시트 생성
        self._create_summary_sheet(wb)
        
        # 2. 각 PID별 상세 정보 시트 생성
        for pid in self.validation_results.keys():
            self._create_pid_detail_sheet(wb, pid)
        
        # 기본 시트 제거
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.save(output_path)
        logger.info(f"통합 리포트 저장 완료: {output_path}")
        print(f"통합 리포트 저장: {output_path}")
    
    def _create_summary_sheet(self, wb):
        """요약 시트 생성"""
        ws = wb.active
        ws.title = "검증 요약"
        
        # 헤더
        headers = ['PID', '기간', '전체 상태', 'PASS', 'FAIL', 'WARNING', 'N/A', '상세 보기']
        ws.append(headers)
        
        # 스타일 정의
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # 헤더 스타일 적용
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 데이터 추가
        row_idx = 2
        for pid, date_ranges in sorted(self.validation_results.items()):
            for date_range, result_info in sorted(date_ranges.items()):
                summary = result_info['summary']
                overall_status = result_info['overall_status']
                
                ws.cell(row=row_idx, column=1, value=pid)
                ws.cell(row=row_idx, column=2, value=date_range)
                
                # 전체 상태
                status_cell = ws.cell(row=row_idx, column=3, value=overall_status)
                if overall_status == 'PASS':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(color="006100", bold=True)
                else:
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(color="9C0006", bold=True)
                
                # 개수
                ws.cell(row=row_idx, column=4, value=summary['PASS'])
                ws.cell(row=row_idx, column=5, value=summary['FAIL'])
                ws.cell(row=row_idx, column=6, value=summary['WARNING'])
                ws.cell(row=row_idx, column=7, value=summary['N/A'])
                
                # 하이퍼링크 (같은 파일 내 시트로 연결)
                detail_sheet_name = f"PID_{pid}"
                link_cell = ws.cell(row=row_idx, column=8, value="자세히 보기")
                link_cell.hyperlink = f"#{detail_sheet_name}!A1"
                link_cell.font = Font(color="0000FF", underline="single")
                link_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 테두리 적용
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).border = thin_border
                    ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
                
                row_idx += 1
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15
    
    def _create_pid_detail_sheet(self, wb, pid):
        """PID별 상세 정보 시트 생성"""
        ws = wb.create_sheet(title=f"PID_{pid}")
        
        # 제목
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"PID {pid} - 검증 상세 결과"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 3
        
        # 각 기간별 결과 표시
        for date_range, result_info in sorted(self.validation_results[pid].items()):
            # 기간 헤더
            ws.merge_cells(f'A{current_row}:G{current_row}')
            period_cell = ws[f'A{current_row}']
            period_cell.value = f"기간: {date_range}"
            period_cell.font = Font(size=12, bold=True)
            period_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            period_cell.alignment = Alignment(horizontal='left', vertical='center')
            period_cell.border = thin_border
            current_row += 1
            
            # 결과 파일 읽기
            result_file = result_info['file_path']
            if os.path.exists(result_file):
                try:
                    result_df = pd.read_excel(result_file, engine='openpyxl')
                    
                    # FAIL 항목만 필터링
                    fail_df = result_df[result_df['Status'] == 'FAIL'].copy()
                    
                    if len(fail_df) > 0:
                        # FAIL 항목 표시
                        fail_header = ws[f'A{current_row}']
                        fail_header.value = f"FAIL 항목 ({len(fail_df)}건)"
                        fail_header.font = Font(bold=True, color="9C0006")
                        current_row += 1
                        
                        # 테이블 헤더
                        table_headers = ['ID', 'Column', 'Check', 'Criteria', 'Fail_Count', 'Details', '원본 파일']
                        for col_idx, header in enumerate(table_headers, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=header)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        current_row += 1
                        
                        # FAIL 데이터 추가
                        for _, row in fail_df.iterrows():
                            ws.cell(row=current_row, column=1, value=row['ID'])
                            ws.cell(row=current_row, column=2, value=row['Column'])
                            ws.cell(row=current_row, column=3, value=row['Check'])
                            ws.cell(row=current_row, column=4, value=row['Criteria'])
                            ws.cell(row=current_row, column=5, value=row['Fail_Count'])
                            ws.cell(row=current_row, column=6, value=row['Details'])
                            
                            # 원본 파일 링크 (외부 파일)
                            source_file = result_file.replace('bms_validation_', '').replace('.xlsx', '.csv')
                            if os.path.exists(source_file):
                                link_cell = ws.cell(row=current_row, column=7, value="원본 보기")
                                # 외부 파일 링크는 절대 경로 사용
                                link_cell.hyperlink = source_file
                                link_cell.font = Font(color="0000FF", underline="single")
                            else:
                                ws.cell(row=current_row, column=7, value="N/A")
                            
                            # 테두리 및 정렬
                            for col in range(1, 8):
                                ws.cell(row=current_row, column=col).border = thin_border
                                ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            
                            current_row += 1
                    else:
                        # FAIL 없음
                        no_fail_cell = ws[f'A{current_row}']
                        no_fail_cell.value = "✓ 모든 항목 PASS"
                        no_fail_cell.font = Font(color="006100", bold=True)
                        current_row += 1
                    
                except Exception as e:
                    logger.error(f"결과 파일 읽기 실패: {result_file} - {e}")
                    error_cell = ws[f'A{current_row}']
                    error_cell.value = f"오류: 결과 파일 읽기 실패 - {str(e)}"
                    error_cell.font = Font(color="FF0000")
                    current_row += 1
            else:
                missing_cell = ws[f'A{current_row}']
                missing_cell.value = "결과 파일 없음"
                missing_cell.font = Font(color="FF0000")
                current_row += 1
            
            current_row += 2  # 간격
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 15


# 테스트용 메인
if __name__ == "__main__":
    # 예시 사용법
    generator = IntegratedReportGenerator("2026-02-10")
    
    generator.add_validation_result(
        pid="2112",
        date_range="2026-01-09_to_2026-01-15",
        result_file_path="downloads/2026-02-10/2112/2026-01-09_to_2026-01-15/bms_validation_MACRIOT_2112.xlsx",
        status_summary={'PASS': 30, 'FAIL': 5, 'WARNING': 2, 'N/A': 0}
    )
    
    generator.generate_integrated_report("downloads/2026-02-10/통합_검증_결과_2026-02-10.xlsx")
