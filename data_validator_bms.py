"""
BMS 데이터 검증 모듈
- 37개 항목의 BMS 데이터 검증 규칙 구현
- S3의 vehicle_master.json에서 차종 정보 로드 (serial_conn_cnt, module_temp_cnt)
- fleet/vehicle_model 컬럼으로 차종 식별 → JSON 매칭
- Fleet 기반 검증: VWGKALRT의 경우 acc_dchg_ah/wh 음수 범위, em_speed_kmh ≤ 255
- VWGKALRT ignit_status 예외: 16/19/20번 항목은 ignit_status=0 행만 검증
- 차종별 N/A: Bolt 2017 → 10-1~3, Ray → 22~25
- 충전 검증: 37-1(Fleet별 충전 검출), 37-2(급속/완속 타입 인식), 35-1(chg*relay)
- unix_time 주기: 3s < interval < 60s만 위반, ≥60s는 다른 key cycle로 제외
- 검증 결과를 xlsx 보고서로 생성
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import json
import boto3
from logger import logger
from typing import Optional


class BMSDataValidator:
    """BMS 데이터 검증 클래스"""
    
    # UI에서 주입되는 AWS 인증 정보 (prdmrt_ui_pyqt.py에서 설정)
    AWS_ACCESS_KEY_ID: Optional[str] = None
    AWS_SECRET_ACCESS_KEY: Optional[str] = None
    AWS_REGION: str = 'ap-northeast-2'
    
    def __init__(self):
        self.df: Optional[pd.DataFrame] = None
        self.results: list = []
        self.failed_rows: dict = {}  # 실패한 행 정보 저장 {item_id: [row_indices]}
        self.vehicle_master: Optional[list] = None  # 차종 정보 리스트 (JSON에서 로드)
        self.fleet: Optional[str] = None           # fleet 컬럼값 (예: 'VWGKALRT')
        self.vehicle_model: Optional[str] = None   # vehicle_model 컬럼값 (예: 'q4-etron')
        self.model_name: Optional[str] = None      # JSON model_name (예: 'Q4-ETRON')
        self.model_year: Optional[str] = None      # JSON model_year (예: '2017')
        self.model_trim: Optional[str] = None      # JSON model_trim (예: 'LONG RANGE')
        self.cell_count: int = 100       # serial_conn_cnt (기본값)
        self.module_count: int = 31      # module_temp_cnt (기본값)
        self._collection_window_mask: Optional[pd.Series] = None  # 16/19/20 수집구간 캐시
        
    # 클래스 레벨 캐시: S3에서 한 번만 로드하면 모든 인스턴스가 공유
    _vehicle_master_cache: Optional[list] = None
    
    def _load_vehicle_master(self):
        """S3에서 차종 정보 JSON 로드
        JSON 구조: list of dict [{model_name, serial_conn_cnt, module_temp_cnt, ...}]
        - 캐시: 한 번 로드하면 클래스 변수에 저장 → 여러 파일 검증 시 S3 재호출 안 함
        - list 형태 그대로 유지 (dict 변환 없음)
        """
        # 이미 캐시된 데이터가 있으면 재사용
        if BMSDataValidator._vehicle_master_cache is not None:
            self.vehicle_master = BMSDataValidator._vehicle_master_cache
            logger.info(f"차종 정보 캐시 사용: {len(self.vehicle_master)}개 모델")
            return
        
        try:
            s3 = boto3.client(
                's3',
                aws_access_key_id=self.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=self.AWS_SECRET_ACCESS_KEY,
                region_name=self.AWS_REGION
            )
            response = s3.get_object(
                Bucket='eplat-validation-monitor',
                Key='tools/download&review/vehicle_master.json'
            )
            raw = json.loads(response['Body'].read().decode('utf-8'))
            # list 형태 그대로 유지
            self.vehicle_master = raw if isinstance(raw, list) else []
            BMSDataValidator._vehicle_master_cache = self.vehicle_master
            logger.info(f"차종 정보 로드 완료: {len(self.vehicle_master)}개 모델")
        except Exception as e:
            logger.warning(f"차종 정보 로드 실패: {e}")
            self.vehicle_master = []
    
    def _get_vehicle_info(self):
        """데이터에서 fleet/vehicle_model 추출 후 JSON에서 셀/모듈 개수 조회
        
        데이터 컬럼:
          - fleet        : 'VWGKALRT' 등 Fleet 코드 (직접 값)
          - vehicle_model: 'q4-etron' 등 차종명 (소문자, 하이픈 포함 가능)
        
        JSON (S3):
          - model_name   : 'Q4-ETRON' 등 (대문자) → vehicle_model과 대소문자 비교
          - serial_conn_cnt: 셀 직렬 개수 (문자열)
          - module_temp_cnt: 모듈 온도 센서 개수 (문자열)
        """
        if self.vehicle_master is None:
            self._load_vehicle_master()
        
        # ── fleet 컬럼에서 Fleet 코드 추출 ──────────────────────────────
        if 'fleet' in self.df.columns and len(self.df) > 0:
            val = self.df['fleet'].dropna()
            self.fleet = str(val.iloc[0]).strip() if len(val) > 0 else None
        else:
            self.fleet = None
            logger.warning("fleet 컬럼 없음")
        
        # ── vehicle_model 컬럼에서 차종명 추출 ──────────────────────────
        if 'vehicle_model' in self.df.columns and len(self.df) > 0:
            val = self.df['vehicle_model'].dropna()
            self.vehicle_model = str(val.iloc[0]).strip() if len(val) > 0 else None
        else:
            self.vehicle_model = None
            logger.warning("vehicle_model 컬럼 없음")
        
        # ── 데이터에서 model_year / model_trim 추출 (컬럼이 있는 경우) ───
        data_model_year = None
        data_model_trim = None
        if 'model_year' in self.df.columns and len(self.df) > 0:
            val = self.df['model_year'].dropna()
            data_model_year = str(val.iloc[0]).strip() if len(val) > 0 else None
        else:
            logger.info("데이터에 model_year 컬럼 없음 → JSON 매칭 시 model_name만 사용")
        if 'model_trim' in self.df.columns and len(self.df) > 0:
            val = self.df['model_trim'].dropna()
            data_model_trim = str(val.iloc[0]).strip() if len(val) > 0 else None
        else:
            logger.info("데이터에 model_trim 컬럼 없음 → JSON 매칭 시 model_name만 사용")
        
        # ── JSON에서 serial_conn_cnt / module_temp_cnt 조회 ─────────────
        # 3단계 매칭: model_name + model_year + model_trim (가장 구체적 → 점진 완화)
        matched = None
        if self.vehicle_master and self.vehicle_model:
            lookup_name = self.vehicle_model.upper()
            
            # 1단계: model_name + model_year + model_trim 완전 일치
            if data_model_year and data_model_trim:
                for item in self.vehicle_master:
                    if (str(item.get('model_name', '')).upper() == lookup_name and
                        str(item.get('model_year', '')).strip() == data_model_year and
                        str(item.get('model_trim', '')).strip().upper() == data_model_trim.upper()):
                        matched = item
                        break
            
            # 2단계: model_name + model_year 일치 (trim 무시)
            if not matched and data_model_year:
                for item in self.vehicle_master:
                    if (str(item.get('model_name', '')).upper() == lookup_name and
                        str(item.get('model_year', '')).strip() == data_model_year):
                        matched = item
                        break
            
            # 3단계: model_name만 일치 (기존 호환)
            if not matched:
                name_matches = [item for item in self.vehicle_master
                                if str(item.get('model_name', '')).upper() == lookup_name]
                if len(name_matches) > 1:
                    logger.warning(
                        f"JSON에 {lookup_name} 모델이 {len(name_matches)}건 존재 "
                        f"(model_year/model_trim 구분 불가 → 첫 번째 항목 사용)"
                    )
                if name_matches:
                    matched = name_matches[0]
        
        if matched:
            try:
                self.cell_count = int(matched.get('serial_conn_cnt', 100))
            except (ValueError, TypeError):
                self.cell_count = 100
            try:
                self.module_count = int(matched.get('module_temp_cnt', 31))
            except (ValueError, TypeError):
                self.module_count = 31
            self.model_name = str(matched.get('model_name', '')).upper()
            self.model_year = str(matched.get('model_year', '')) if matched.get('model_year') else None
            self.model_trim = str(matched.get('model_trim', '')) if matched.get('model_trim') else None
            logger.info(
                f"차종 매칭 성공 | fleet={self.fleet}, "
                f"vehicle_model={self.vehicle_model}, "
                f"model_name={self.model_name}, model_year={self.model_year}, "
                f"model_trim={self.model_trim}, "
                f"serial_conn_cnt={self.cell_count}, "
                f"module_temp_cnt={self.module_count}"
            )
        else:
            self.cell_count = 100
            self.module_count = 31
            self.model_name = None
            self.model_year = None
            self.model_trim = None
            logger.warning(
                f"차종 정보 없음 (vehicle_model={self.vehicle_model}), 기본값 사용"
            )
    
    def _get_used_columns(self):
        """검증에 사용하는 컬럼 목록 반환 (CSV 로드 시 필요 컬럼만 선별)"""
        # 기본 검증 컬럼
        base_cols = [
            'unix_time', 'ignit_status', 'chg_conr_status_list', 'em_speed_kmh',
            'pack_curr', 'pack_volt', 'main_relay_status', 'soc_display_rate',
            'soc_rate', 'mile_km', 'cell_min_volt', 'cell_max_volt', 'cell_volt_dev',
            'cell_min_volt_no', 'cell_max_volt_no', 'module_min_temp', 'module_max_temp',
            'module_avg_temp', 'oper_second', 'acc_chg_ah', 'acc_dchg_ah',
            'acc_chg_wh', 'acc_dchg_wh', 'pack_pwr', 'ir', 'soh_rate',
            'fleet', 'vehicle_model', 'model_year', 'model_trim',
        ]
        # 동적 컬럼: cell_voltage_1~N, battery_module_N_temperature
        cell_cols = [f'cell_voltage_{i}' for i in range(1, 201)]
        module_cols = [f'battery_module_{i}_temperature' for i in range(1, 51)]
        return base_cols + cell_cols + module_cols
    
    def validate_file(self, file_path):
        """
        파일 검증 메인 함수
        :param file_path: CSV 또는 Parquet 파일 경로
        :return: 검증 결과 리스트
        """
        logger.info(f"파일 검증 시작: {file_path}")
        
        # 파일 로드 (필요 컬럼만 선별하여 메모리/속도 최적화)
        used_cols = self._get_used_columns()
        if file_path.endswith('.csv'):
            # 먼저 헤더만 읽어서 실제 존재하는 컬럼과 교집합
            all_cols = pd.read_csv(file_path, nrows=0).columns.tolist()
            load_cols = [c for c in used_cols if c in all_cols]
            self.df = pd.read_csv(file_path, usecols=load_cols, low_memory=False)
        elif file_path.endswith('.parquet'):
            import pyarrow.parquet as pq
            schema_cols = pq.read_schema(file_path).names
            load_cols = [c for c in used_cols if c in schema_cols]
            self.df = pd.read_parquet(file_path, columns=load_cols)
        else:
            raise ValueError("지원하지 않는 파일 형식")
        
        logger.info(f"데이터 로드: {len(self.df)}행, {len(self.df.columns)}열")
        
        # 차종 정보 추출
        self._get_vehicle_info()
        
        # 모든 검증 수행
        self.results = []
        self.failed_rows = {}
        self._collection_window_mask = None  # 수집구간 캐시 초기화
        self._validate_all()
        
        logger.info(f"검증 완료: 총 {len(self.results)}개 항목")

        return self.results
    
    def _validate_all(self):
        """모든 검증 항목 실행"""
        # 1. unix_time
        self._validate_1_unix_time()
        
        # 2. ignit_status
        self._validate_2_ignit_status()
        
        # 3. chg_conr_status_list
        self._validate_3_chg_conr_status_list()
        
        # 4. em_speed_kmh
        self._validate_4_em_speed_kmh()
        
        # 5. pack_curr
        self._validate_5_pack_curr()
        
        # 6. pack_volt
        self._validate_6_pack_volt()
        
        # 7. main_relay_status
        self._validate_7_main_relay_status()
        
        # 8. soc_display_rate
        self._validate_8_soc_display_rate()
        
        # 9. soc_rate
        self._validate_9_soc_rate()
        
        # 10. mile_km
        self._validate_10_mile_km()
        
        # 11. cell_min_volt
        self._validate_11_cell_min_volt()
        
        # 12. cell_max_volt
        self._validate_12_cell_max_volt()
        
        # 13. cell_volt_dev
        self._validate_13_cell_volt_dev()
        
        # 14. cell_min_volt_no
        self._validate_14_cell_min_volt_no()
        
        # 15. cell_max_volt_no
        self._validate_15_cell_max_volt_no()
        
        # 16. cell_voltage_1~N
        self._validate_16_cell_voltages()
        
        # 17. module_min_temp
        self._validate_17_module_min_temp()
        
        # 18. module_max_temp
        self._validate_18_module_max_temp()
        
        # 19. module_avg_temp
        self._validate_19_module_avg_temp()
        
        # 20. battery_module_N_temperature
        self._validate_20_battery_modules()
        
        # 21. oper_second
        self._validate_21_oper_second()
        
        # 22. acc_chg_ah
        self._validate_22_acc_chg_ah()
        
        # 23. acc_dchg_ah
        self._validate_23_acc_dchg_ah()
        
        # 24. acc_chg_wh
        self._validate_24_acc_chg_wh()
        
        # 25. acc_dchg_wh
        self._validate_25_acc_dchg_wh()
        
        # 26. pack_pwr
        self._validate_26_pack_pwr()
        
        # 27. ir
        self._validate_27_ir()
        
        # 28. 충전 전류 소요시간
        self._validate_28_charging_time()
        
        # 29. 데이터 시작/종료 status
        self._validate_29_start_end_status()
        
        # 30. OBD Sleep Latency
        self._validate_30_sleep_latency()
        
        # 31. soh_rate
        self._validate_31_soh_rate()
        
        # 32. IGN * vehicle speed
        self._validate_32_ign_speed()
        
        # 33. IGN * main relay
        self._validate_33_ign_relay()
        
        # 34. chg * main relay
        self._validate_34_chg_relay()
        
        # 35. chg 충전 조건 검증
        self._validate_35_chg_condition()
    
    def _add_result(self, item_id, column, check, criteria, status, fail_count, details, failed_indices=None):
        """
        검증 결과 추가
        :param failed_indices: 실패한 행 인덱스 리스트
        """
        # 에러 상세: 첫 번째 실패 행의 데이터 샘플 생성
        error_sample = ''
        if failed_indices is not None and len(failed_indices) > 0 and self.df is not None:
            try:
                first_idx = failed_indices[0]
                if first_idx in self.df.index:
                    row = self.df.loc[first_idx]
                    # 관련 컬럼 값을 샘플로 추출
                    sample_cols = self._get_error_sample_cols(column, item_id)
                    sample_parts = [f'idx={first_idx}']
                    if 'signal_kst_ts' in self.df.columns:
                        ts_val = row.get('signal_kst_ts', None)
                        if ts_val is not None and pd.notna(ts_val):
                            try:
                                import re as _re
                                _m = _re.search(r'(\d{1,2}:\d{2})', str(ts_val))
                                if _m:
                                    sample_parts.append(f'time={_m.group(1)}')
                            except Exception:
                                pass
                    for sc in sample_cols:
                        if sc in self.df.columns:
                            val = row[sc]
                            if pd.notna(val):
                                if isinstance(val, float):
                                    sample_parts.append(f'{sc}={val:.4g}')
                                else:
                                    sample_parts.append(f'{sc}={val}')
                            else:
                                sample_parts.append(f'{sc}=NaN')
                    error_sample = ' | 첫 에러: ' + ', '.join(sample_parts)
            except Exception:
                pass
        
        self.results.append({
            'ID': item_id,
            'Column': column,
            'Check': check,
            'Criteria': criteria,
            'Status': status,
            'Fail_Count': int(fail_count),
            'Details': details + error_sample
        })
        
        # 실패한 행 정보 저장
        if failed_indices is not None and len(failed_indices) > 0:
            self.failed_rows[item_id] = failed_indices
    
    def _get_error_sample_cols(self, column, item_id):
        """에러 샘플에 표시할 관련 컬럼 목록 반환"""
        # 항목별 관련 컬럼 매핑
        base_map = {
            'unix_time': ['unix_time'],
            'ignit_status': ['ignit_status'],
            'chg_conr_status_list': ['chg_conr_status_list'],
            'em_speed_kmh': ['em_speed_kmh', 'ignit_status'],
            'pack_curr': ['pack_curr'],
            'pack_volt': ['pack_volt'],
            'main_relay_status': ['main_relay_status'],
            'soc_display_rate': ['soc_display_rate'],
            'soc_rate': ['soc_rate'],
            'mile_km': ['mile_km'],
            'cell_min_volt': ['cell_min_volt'],
            'cell_max_volt': ['cell_max_volt'],
            'cell_volt_dev': ['cell_volt_dev', 'cell_max_volt', 'cell_min_volt'],
            'cell_min_volt_no': ['cell_min_volt_no'],
            'cell_max_volt_no': ['cell_max_volt_no'],
            'module_min_temp': ['module_min_temp'],
            'module_max_temp': ['module_max_temp'],
            'module_avg_temp': ['module_avg_temp'],
            'oper_second': ['oper_second'],
            'soh_rate': ['soh_rate'],
            'pack_pwr': ['pack_pwr', 'pack_curr', 'pack_volt'],
            'ir': ['ir'],
        }
        # 복합 항목
        if 'cell_voltage' in column:
            return ['ignit_status', 'chg_conr_status_list', 'cell_voltage_1']
        if 'battery_module' in column:
            return ['ignit_status', 'chg_conr_status_list', 'battery_module_1_temperature']
        if 'acc_chg_ah' in column:
            return ['acc_chg_ah']
        if 'acc_dchg_ah' in column:
            return ['acc_dchg_ah']
        if 'acc_chg_wh' in column:
            return ['acc_chg_wh']
        if 'acc_dchg_wh' in column:
            return ['acc_dchg_wh']
        if 'IGN * vehicle speed' in column:
            return ['ignit_status', 'em_speed_kmh']
        if 'IGN * main relay' in column:
            return ['ignit_status', 'main_relay_status']
        if 'chg * main relay' in column:
            return ['chg_conr_status_list', 'main_relay_status']
        if '충전 전류' in column:
            return ['chg_conr_status_list', 'pack_curr', 'unix_time']
        if 'Sleep Latency' in column:
            return ['ignit_status', 'chg_conr_status_list', 'unix_time']
        if '시작/종료' in column:
            return ['ignit_status', 'chg_conr_status_list', 'main_relay_status']
        
        return base_map.get(column, [column])
    
    def _check_column_exists(self, col):
        """컬럼 존재 여부 확인"""
        return col in self.df.columns
    
    def _get_interval_violations(self, col, max_interval=20):
        """
        수집 주기 위반 검사 (numpy 벡터 연산)
        :param col: 컬럼명
        :param max_interval: 최대 간격
        :return: 위반 횟수, 위반 인덱스 리스트
        """
        if not self._check_column_exists(col):
            return 0, []
        
        # NULL이 아닌 값들의 인덱스 (numpy 배열)
        valid_indices = self.df.index[self.df[col].notna()].values
        
        if len(valid_indices) < 2:
            return 0, []
        
        # numpy 벡터 연산으로 간격 계산
        intervals = np.diff(valid_indices)
        violation_mask = intervals > max_interval
        violation_positions = np.where(violation_mask)[0]
        fail_indices = valid_indices[violation_positions + 1].tolist()
        
        return len(fail_indices), fail_indices
    
    def _get_collection_window_mask(self):
        """16/19/20번 항목의 수집 구간 마스크 생성
        
        이벤트(key cycle) 내에서 '수집 활성' 상태를 추적:
        - IGN 1→0 전환 시 수집 활성화 (IGN OFF 구간)
        - IGN이 다시 1이 되면(재시동) 수집 비활성화
        - 충전 중(1|0 or 0|1)이면 수집 활성화
        - 충전 종료(1|0/0|1→0|0) 시 수집 유지 (IGN=0 상태이므로 sleep까지)
        - IGN=0이고 충전 안 중이면 이전 수집 상태 유지 (sleep 대기 구간)
        """
        if self._collection_window_mask is not None:
            return self._collection_window_mask
        
        mask = pd.Series(False, index=self.df.index)
        events = self._split_into_events()
        
        for evt_df in events:
            if len(evt_df) < 2:
                continue
            
            has_ignit = 'ignit_status' in evt_df.columns
            has_chg = 'chg_conr_status_list' in evt_df.columns
            
            collecting = False
            prev_ignit = None
            prev_chg_str = None
            
            for idx in evt_df.index:
                curr_ignit = evt_df.at[idx, 'ignit_status'] if has_ignit else None
                curr_chg_str = str(evt_df.at[idx, 'chg_conr_status_list']) if has_chg else '0|0'
                curr_charging = curr_chg_str in ('1|0', '0|1')
                
                # 재시동(IGN=1) 시 수집 비활성화 (충전 중이 아닐 때만)
                if collecting and has_ignit and curr_ignit == 1 and not curr_charging:
                    collecting = False
                
                if not collecting:
                    # (a) IGN 1→0 전환 → 수집 시작
                    if has_ignit and prev_ignit == 1 and curr_ignit == 0:
                        collecting = True
                    # (b) 충전 시작 → 수집 시작
                    if curr_charging:
                        collecting = True
                    # (c) 충전 종료(1|0/0|1→0|0) → 수집 시작 (sleep 대기)
                    if has_chg and prev_chg_str in ('1|0', '0|1') and curr_chg_str == '0|0':
                        collecting = True
                
                if collecting:
                    mask.at[idx] = True
                
                prev_ignit = curr_ignit
                prev_chg_str = curr_chg_str
        
        self._collection_window_mask = mask
        return mask
    
    def _get_interval_violations_in_windows(self, col, max_interval=20):
        """수집 구간(collection window) 내에서만 수집 주기 위반 검사
        
        수집 구간 외의 null 간격은 정상적 미수집으로 무시.
        같은 연속 수집 구간 내에서만 interval 체크.
        """
        if not self._check_column_exists(col):
            return 0, []
        
        collection_mask = self._get_collection_window_mask()
        
        # 수집 구간 내에서 non-null인 행의 인덱스
        in_window_valid = collection_mask & self.df[col].notna()
        valid_indices = self.df.index[in_window_valid].values
        
        if len(valid_indices) < 2:
            return 0, []
        
        # 연속 수집 구간(segment) 라벨 생성
        mask_int = collection_mask.astype(int)
        starts = (mask_int.diff().fillna(mask_int.iloc[0]) == 1)
        segments = starts.cumsum().values.copy()
        segments[~collection_mask.values] = 0  # 비수집 구간은 0
        
        # 간격 계산 및 위반 검사 (같은 segment 내에서만)
        intervals = np.diff(valid_indices)
        fail_indices = []
        for i in range(len(intervals)):
            if intervals[i] > max_interval:
                seg_from = segments[valid_indices[i]]
                seg_to = segments[valid_indices[i + 1]]
                if seg_from == seg_to and seg_from > 0:
                    fail_indices.append(valid_indices[i + 1])
        
        return len(fail_indices), fail_indices
    
    # ==================== 검증 함수들 ====================
    
    def _validate_1_unix_time(self):
        """1. unix_time 검증"""
        col = 'unix_time'
        
        if not self._check_column_exists(col):
            self._add_result('1-1', col, '수집 주기', '3초 초과 ~ 60초 미만 (≥60초는 다른 key cycle)', 'N/A', 0, '컬럼 없음')
            self._add_result('1-2', col, 'null 체크', '값이 존재함', 'N/A', 0, '컬럼 없음')
            self._add_result('1-3', col, '값 유효성', '1,420,070,400 ≤ unix_time ≤ 2,524,608,000', 'N/A', 0, '컬럼 없음')
            return
        
        # 1-1: 수집 주기 (3초 초과 ~ 60초 미만만 위반, ≥60초는 다른 key cycle로 제외)
        valid_times = self.df[self.df[col].notna()][col].values
        if len(valid_times) >= 2:
            intervals = np.diff(valid_times)
            violations = np.where((intervals > 3) & (intervals < 60))[0]
            fail_indices = [i + 1 for i in violations]  # diff의 인덱스를 원본 인덱스로 변환
            status = 'PASS' if len(violations) == 0 else 'FAIL'
            self._add_result('1-1', col, '수집 주기', '3초 초과 ~ 60초 미만 (≥60초는 다른 key cycle)', status, len(violations), 
                           f'{len(violations)}건 위반' if violations.size > 0 else '정상',
                           fail_indices)
        else:
            self._add_result('1-1', col, '수집 주기', '3초 이내', 'WARNING', 0, '데이터 부족')
        
        # 1-2: null 체크
        null_count = self.df[col].isna().sum()
        null_indices = self.df[self.df[col].isna()].index.tolist()
        status = 'PASS' if null_count == 0 else 'FAIL'
        self._add_result('1-2', col, 'null 체크', '값이 존재함', status, null_count, 
                       f'{null_count}개 null' if null_count > 0 else '정상',
                       null_indices)
        
        # 1-3: 값 유효성 (1,420,070,400 ≤ unix_time ≤ 2,524,608,000)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < 1420070400) | (valid_data[col] > 2524608000)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('1-3', col, '값 유효성', '1,420,070,400 ≤ unix_time ≤ 2,524,608,000', 
                       status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_2_ignit_status(self):
        """2. ignit_status 검증"""
        col = 'ignit_status'
        
        if not self._check_column_exists(col):
            self._add_result('2-1', col, 'null 체크', '값이 존재함', 'N/A', 0, '컬럼 없음')
            self._add_result('2-2', col, '값 유효성', '0 or 1', 'N/A', 0, '컬럼 없음')
            return
        
        # 2-1: null 체크
        null_count = self.df[col].isna().sum()
        null_indices = self.df[self.df[col].isna()].index.tolist()
        status = 'PASS' if null_count == 0 else 'FAIL'
        self._add_result('2-1', col, 'null 체크', '값이 존재함', status, null_count,
                       f'{null_count}개 null' if null_count > 0 else '정상',
                       null_indices)
        
        # 2-2: 값 유효성 (0 or 1)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = ~valid_data[col].isin([0, 1])
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('2-2', col, '값 유효성', '0 or 1', status, invalid_count,
                       f'{invalid_count}건 유효하지 않은 값' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_3_chg_conr_status_list(self):
        """3. chg_conr_status_list 검증"""
        col = 'chg_conr_status_list'
        
        if not self._check_column_exists(col):
            self._add_result('3-1', col, 'null 체크', '값이 존재함', 'N/A', 0, '컬럼 없음')
            self._add_result('3-2', col, '값 유효성', '0|0 or 0|1 or 1|0', 'N/A', 0, '컬럼 없음')
            return
        
        # 3-1: null 체크
        null_count = self.df[col].isna().sum()
        null_indices = self.df[self.df[col].isna()].index.tolist()
        status = 'PASS' if null_count == 0 else 'FAIL'
        self._add_result('3-1', col, 'null 체크', '값이 존재함', status, null_count,
                       f'{null_count}개 null' if null_count > 0 else '정상',
                       null_indices)
        
        # 3-2: 값 유효성 (0|0, 0|1, 1|0)
        valid_data = self.df[self.df[col].notna()]
        valid_values = ['0|0', '0|1', '1|0']
        invalid_mask = ~valid_data[col].astype(str).isin(valid_values)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('3-2', col, '값 유효성', '0|0 or 0|1 or 1|0', status, invalid_count,
                       f'{invalid_count}건 유효하지 않은 값' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_4_em_speed_kmh(self):
        """4. em_speed_kmh 검증 (Fleet별 기준)"""
        col = 'em_speed_kmh'
        
        # Fleet별 최대 속도 설정
        max_speed = 255.0 if self.fleet == 'VWGKALRT' else 250.0
        criteria_text = f'0.0 ≤ em_speed_kmh ≤ {max_speed}'
        
        if not self._check_column_exists(col):
            self._add_result('4-1', col, 'null 체크', '값이 존재함', 'N/A', 0, '컬럼 없음')
            self._add_result('4-2', col, '값 유효성', criteria_text, 'N/A', 0, '컬럼 없음')
            return
        
        # 4-1: null 체크
        null_count = self.df[col].isna().sum()
        null_indices = self.df[self.df[col].isna()].index.tolist()
        status = 'PASS' if null_count == 0 else 'FAIL'
        self._add_result('4-1', col, 'null 체크', '값이 존재함', status, null_count,
                       f'{null_count}개 null' if null_count > 0 else '정상',
                       null_indices)
        
        # 4-2: 값 유효성 (Fleet별 범위)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < 0.0) | (valid_data[col] > max_speed)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('4-2', col, '값 유효성', criteria_text, status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_5_pack_curr(self):
        """5. pack_curr 검증"""
        col = 'pack_curr'
        
        if not self._check_column_exists(col):
            self._add_result('5-1', col, 'null 체크', '값이 존재함', 'N/A', 0, '컬럼 없음')
            self._add_result('5-2', col, '값 유효성', '-360.0 ≤ pack_curr ≤ 750.0', 'N/A', 0, '컬럼 없음')
            return
        
        # 5-1: null 체크
        null_count = self.df[col].isna().sum()
        null_indices = self.df[self.df[col].isna()].index.tolist()
        status = 'PASS' if null_count == 0 else 'FAIL'
        self._add_result('5-1', col, 'null 체크', '값이 존재함', status, null_count,
                       f'{null_count}개 null' if null_count > 0 else '정상',
                       null_indices)
        
        # 5-2: 값 유효성 (-360.0 ~ 750.0)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < -360.0) | (valid_data[col] > 750.0)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('5-2', col, '값 유효성', '-360.0 ≤ pack_curr ≤ 750.0', status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_6_pack_volt(self):
        """6. pack_volt 검증 (셀 전압 × 셀 개수 기반 동적 계산)"""
        col = 'pack_volt'
        
        # 셀 전압 범위: 2.6V ~ 4.5V
        cell_volt_min = 2.6
        cell_volt_max = 4.5
        
        # pack_volt 범위 = 셀 전압 범위 × 셀 개수
        pack_volt_min = cell_volt_min * self.cell_count
        pack_volt_max = cell_volt_max * self.cell_count
        
        criteria_text = f'{pack_volt_min:.1f} ≤ pack_volt ≤ {pack_volt_max:.1f} (셀전압 {cell_volt_min}~{cell_volt_max}V × {self.cell_count}셀)'
        
        if not self._check_column_exists(col):
            self._add_result('6-1', col, 'null 체크', '값이 존재함', 'N/A', 0, '컬럼 없음')
            self._add_result('6-2', col, '값 유효성', criteria_text, 'N/A', 0, '컬럼 없음')
            return
        
        # 6-1: null 체크
        null_count = self.df[col].isna().sum()
        null_indices = self.df[self.df[col].isna()].index.tolist()
        status = 'PASS' if null_count == 0 else 'FAIL'
        self._add_result('6-1', col, 'null 체크', '값이 존재함', status, null_count,
                       f'{null_count}개 null' if null_count > 0 else '정상',
                       null_indices)
        
        # 6-2: 값 유효성 (동적 범위)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < pack_volt_min) | (valid_data[col] > pack_volt_max)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('6-2', col, '값 유효성', criteria_text, status, invalid_count,
                       f'{invalid_count}건 범위 초과 (셀 개수: {self.cell_count})' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_7_main_relay_status(self):
        """7. main_relay_status 검증"""
        col = 'main_relay_status'
        
        if not self._check_column_exists(col):
            self._add_result('7-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('7-2', col, '값 유효성', '0 or 1', 'N/A', 0, '컬럼 없음')
            return
        
        # 7-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('7-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 7-2: 값 유효성 (0 or 1)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = ~valid_data[col].isin([0, 1])
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('7-2', col, '값 유효성', '0 or 1', status, invalid_count,
                       f'{invalid_count}건 유효하지 않은 값' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_8_soc_display_rate(self):
        """8. soc_display_rate 검증"""
        col = 'soc_display_rate'
        
        if not self._check_column_exists(col):
            self._add_result('8-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('8-2', col, '값 유효성', '0 ≤ soc_display_rate ≤ 100.0', 'N/A', 0, '컬럼 없음')
            return
        
        # 8-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('8-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 8-2: 값 유효성 (0 ~ 100)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < 0) | (valid_data[col] > 100.0)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('8-2', col, '값 유효성', '0 ≤ soc_display_rate ≤ 100.0', status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_9_soc_rate(self):
        """9. soc_rate 검증 (시작 SOC 0 허용으로 변경)"""
        col = 'soc_rate'
        
        if not self._check_column_exists(col):
            self._add_result('9-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('9-2', col, '값 유효성', '0.1 ≤ soc_rate ≤ 100.0', 'N/A', 0, '컬럼 없음')
            return
        
        # 9-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('9-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 9-2: 값 유효성 (0.1 ~ 100.0)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < 0.1) | (valid_data[col] > 100.0)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('9-2', col, '값 유효성', '0.1 ≤ soc_rate ≤ 100.0', status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_10_mile_km(self):
        """10. mile_km 검증 (Bolt 2017은 미수집 → N/A)"""
        col = 'mile_km'
        
        # Bolt 2017 차종은 mile_km 미수집 → N/A 처리
        if self.model_name == 'BOLT' and self.model_year == '2017':
            self._add_result('10-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '차종 미수집 (BOLT 2017)')
            self._add_result('10-2', col, '값 유효성', '1 ≤ mile_km ≤ 1,000,000', 'N/A', 0, '차종 미수집 (BOLT 2017)')
            self._add_result('10-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1', 'N/A', 0, '차종 미수집 (BOLT 2017)')
            return
        
        if not self._check_column_exists(col):
            self._add_result('10-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('10-2', col, '값 유효성', '1 ≤ mile_km ≤ 1,000,000', 'N/A', 0, '컬럼 없음')
            self._add_result('10-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1', 'N/A', 0, '컬럼 없음')
            return
        
        # 10-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('10-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 10-2: 값 유효성 (1 ~ 1,000,000)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < 1) | (valid_data[col] > 1000000)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('10-2', col, '값 유효성', '1 ≤ mile_km ≤ 1,000,000', status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
        
        # 10-3: 값 선형 증가 (차이 ≤ 1)
        valid_values = self.df[self.df[col].notna()][col].values
        if len(valid_values) >= 2:
            diffs = np.diff(valid_values)
            violations = np.where(diffs > 1)[0]
            fail_indices = [i + 1 for i in violations]
            status = 'PASS' if len(violations) == 0 else 'FAIL'
            self._add_result('10-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1',
                           status, len(violations),
                           f'{len(violations)}건 위반' if len(violations) > 0 else '정상',
                           fail_indices)
        else:
            self._add_result('10-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1',
                           'WARNING', 0, '데이터 부족')
    
    def _validate_11_cell_min_volt(self):
        """11. cell_min_volt 검증"""
        col = 'cell_min_volt'
        
        if not self._check_column_exists(col):
            self._add_result('11-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('11-2', col, '값 유효성', '2.6 ≤ cell_min_volt ≤ 4.5', 'N/A', 0, '컬럼 없음')
            return
        
        # 11-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('11-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 11-2: 값 유효성 (2.6 ~ 4.5)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < 2.6) | (valid_data[col] > 4.5)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('11-2', col, '값 유효성', '2.6 ≤ cell_min_volt ≤ 4.5', status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_12_cell_max_volt(self):
        """12. cell_max_volt 검증"""
        col = 'cell_max_volt'
        
        if not self._check_column_exists(col):
            self._add_result('12-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('12-2', col, '값 유효성', '2.6 ≤ cell_max_volt ≤ 4.5', 'N/A', 0, '컬럼 없음')
            return
        
        # 12-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('12-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 12-2: 값 유효성 (2.6 ~ 4.5)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < 2.6) | (valid_data[col] > 4.5)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('12-2', col, '값 유효성', '2.6 ≤ cell_max_volt ≤ 4.5', status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_13_cell_volt_dev(self):
        """13. cell_volt_dev 검증"""
        col = 'cell_volt_dev'
        
        if not self._check_column_exists(col):
            self._add_result('13-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('13-2', col, '값 유효성', 'cell_volt_dev = cell_max_volt - cell_min_volt', 'N/A', 0, '컬럼 없음')
            return
        
        # 13-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('13-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 13-2: cell_volt_dev = cell_max_volt - cell_min_volt 검증
        criteria_text = 'cell_volt_dev = cell_max_volt - cell_min_volt'
        if 'cell_max_volt' not in self.df.columns or 'cell_min_volt' not in self.df.columns:
            self._add_result('13-2', col, '값 유효성', criteria_text, 'N/A', 0,
                           'cell_max_volt 또는 cell_min_volt 컬럼 없음')
            return
        
        valid_mask = self.df[col].notna() & self.df['cell_max_volt'].notna() & self.df['cell_min_volt'].notna()
        valid_data = self.df[valid_mask]
        calculated = (valid_data['cell_max_volt'] - valid_data['cell_min_volt']).round(6)
        actual = valid_data[col].round(6)
        # 부동소수점 오차 허용 (±0.001V)
        mismatch_mask = (actual - calculated).abs() > 0.001
        invalid_count = mismatch_mask.sum()
        invalid_indices = valid_data[mismatch_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('13-2', col, '값 유효성', criteria_text, status, invalid_count,
                       f'{invalid_count}건 불일치 (cell_max_volt - cell_min_volt 값과 다름)' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_14_cell_min_volt_no(self):
        """14. cell_min_volt_no 검증"""
        col = 'cell_min_volt_no'
        
        n = self.cell_count
        criteria_n = f'cell_min_volt_no ≤ {n}'
        
        if not self._check_column_exists(col):
            self._add_result('14-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('14-2', col, '값 유효성', criteria_n, 'N/A', 0, '컬럼 없음')
            return
        
        # 14-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('14-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 14-2: 값 유효성 (≤ n)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = valid_data[col] > n
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('14-2', col, '값 유효성', criteria_n, status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_15_cell_max_volt_no(self):
        """15. cell_max_volt_no 검증"""
        col = 'cell_max_volt_no'
        
        n = self.cell_count
        criteria_n = f'cell_max_volt_no ≤ {n}'
        
        if not self._check_column_exists(col):
            self._add_result('15-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('15-2', col, '값 유효성', criteria_n, 'N/A', 0, '컬럼 없음')
            return
        
        # 15-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('15-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 15-2: 값 유효성 (≤ n)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = valid_data[col] > n
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('15-2', col, '값 유효성', criteria_n, status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_16_cell_voltages(self):
        """16. cell_voltage_1~N 검증 (vehicle_master 기반 개수 확인)
        
        수집 구간: IGN 1→0 ~sleep, 충전중(1|0/0|1), 충전종료→0|0 ~sleep
        """
        # cell_voltage로 시작하는 컬럼 찾기
        cell_cols = [col for col in self.df.columns if col.startswith('cell_voltage_')]
        
        expected_count = self.cell_count
        criteria_count = f'{expected_count}개 이상'
        
        if len(cell_cols) == 0:
            self._add_result('16-1', 'cell_voltage_1~N', '수집 주기', '수집 구간 내 ≤20', 'N/A', 0, '컬럼 없음')
            self._add_result('16-2', 'cell_voltage_1~N', '데이터 개수', criteria_count, 'N/A', 0, '컬럼 없음')
            self._add_result('16-3', 'cell_voltage_1~N', '값 유효성(최소)', '2.6 ≤ 개별 셀 전압 최소값 ≤ 4.3', 'N/A', 0, '컬럼 없음')
            self._add_result('16-4', 'cell_voltage_1~N', '값 유효성(최대)', '2.6 ≤ 개별 셀 전압 최대값 ≤ 4.3', 'N/A', 0, '컬럼 없음')
            return
        
        # 수집 구간 마스크 (IGN OFF~sleep, 충전중, 충전종료~sleep)
        collection_mask = self._get_collection_window_mask()
        df_target = self.df[collection_mask]
        window_note = ' (수집구간: IGN OFF→sleep/충전중/충전종료→sleep)'
        
        # 16-1: 수집 주기 (수집 구간 내에서만 검사)
        first_col = cell_cols[0]
        fail_count, fail_indices = self._get_interval_violations_in_windows(first_col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('16-1', 'cell_voltage_1~N', '수집 주기', '수집 구간 내 ≤20',
                       status, fail_count,
                       f'{fail_count}건 위반{window_note}' if fail_count > 0 else f'정상{window_note}',
                       fail_indices)
        
        # 16-2: 데이터 개수 (vehicle_master 기반) - 실제 데이터가 있는 컬럼만 카운트
        # 전체 NaN인 컬럼은 제외 (스키마상 존재하지만 해당 차종에서 미사용)
        active_cell_cols = [c for c in cell_cols if self.df[c].notna().any()]
        actual_count = len(active_cell_cols)
        if actual_count == expected_count:
            status = 'PASS'
            details = f'일치 (기준: {expected_count}개, 실제: {actual_count}개)'
            fail_cnt = 0
        elif actual_count > expected_count:
            status = 'PASS'
            details = f'기준 초과 (기준: {expected_count}개, 실제: {actual_count}개)'
            fail_cnt = 0
        else:
            status = 'FAIL'
            details = f'불일치 (기준: {expected_count}개, 실제: {actual_count}개)'
            fail_cnt = expected_count - actual_count
        self._add_result('16-2', 'cell_voltage_1~N', '데이터 개수', criteria_count,
                       status, fail_cnt, details)
        
        # 16-3: 최소값 유효성 (2.6 ~ 4.3) - 수집 구간만
        if len(df_target) > 0:
            all_min_values = df_target[cell_cols].min(axis=1)
            valid_data = all_min_values[all_min_values.notna()]
            invalid_mask = (valid_data < 2.6) | (valid_data > 4.3)
            invalid_count = invalid_mask.sum()
            invalid_indices = valid_data[invalid_mask].index.tolist()
        else:
            invalid_count = 0
            invalid_indices = []
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('16-3', 'cell_voltage_1~N', '값 유효성(최소)', '2.6 ≤ 개별 셀 전압 최소값 ≤ 4.3',
                       status, invalid_count,
                       f'{invalid_count}건 범위 초과{window_note}' if invalid_count > 0 else f'정상{window_note}',
                       invalid_indices)
        
        # 16-4: 최대값 유효성 (2.6 ~ 4.3) - 수집 구간만
        if len(df_target) > 0:
            all_max_values = df_target[cell_cols].max(axis=1)
            valid_data = all_max_values[all_max_values.notna()]
            invalid_mask = (valid_data < 2.6) | (valid_data > 4.3)
            invalid_count = invalid_mask.sum()
            invalid_indices = valid_data[invalid_mask].index.tolist()
        else:
            invalid_count = 0
            invalid_indices = []
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('16-4', 'cell_voltage_1~N', '값 유효성(최대)', '2.6 ≤ 개별 셀 전압 최대값 ≤ 4.3',
                       status, invalid_count,
                       f'{invalid_count}건 범위 초과{window_note}' if invalid_count > 0 else f'정상{window_note}',
                       invalid_indices)
    
    def _validate_17_module_min_temp(self):
        """17. module_min_temp 검증"""
        col = 'module_min_temp'
        
        if not self._check_column_exists(col):
            self._add_result('17-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('17-2', col, '값 유효성', '-20.0 ≤ module_min_temp ≤ 60.0', 'N/A', 0, '컬럼 없음')
            return
        
        # 17-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('17-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 17-2: 값 유효성 (-20 ~ 60)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < -20.0) | (valid_data[col] > 60.0)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('17-2', col, '값 유효성', '-20.0 ≤ module_min_temp ≤ 60.0', status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_18_module_max_temp(self):
        """18. module_max_temp 검증"""
        col = 'module_max_temp'
        
        if not self._check_column_exists(col):
            self._add_result('18-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('18-2', col, '값 유효성', '-20.0 ≤ module_max_temp ≤ 60.0', 'N/A', 0, '컬럼 없음')
            return
        
        # 18-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('18-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 18-2: 값 유효성 (-20 ~ 60)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < -20.0) | (valid_data[col] > 60.0)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('18-2', col, '값 유효성', '-20.0 ≤ module_max_temp ≤ 60.0', status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_19_module_avg_temp(self):
        """19. module_avg_temp 검증
        
        수집 구간: IGN 1→0 ~sleep, 충전중(1|0/0|1), 충전종료→0|0 ~sleep
        """
        col = 'module_avg_temp'
        
        if not self._check_column_exists(col):
            self._add_result('19-1', col, '수집 주기', '수집 구간 내 ≤20', 'N/A', 0, '컬럼 없음')
            self._add_result('19-2', col, '값 유효성', '-20.0 ≤ module_avg_temp ≤ 60.0', 'N/A', 0, '컬럼 없음')
            return
        
        # 수집 구간 마스크
        collection_mask = self._get_collection_window_mask()
        df_target = self.df[collection_mask]
        window_note = ' (수집구간: IGN OFF→sleep/충전중/충전종료→sleep)'
        
        # 19-1: 수집 주기 (수집 구간 내에서만)
        fail_count, fail_indices = self._get_interval_violations_in_windows(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('19-1', col, '수집 주기', '수집 구간 내 ≤20',
                       status, fail_count,
                       f'{fail_count}건 위반{window_note}' if fail_count > 0 else f'정상{window_note}',
                       fail_indices)
        
        # 19-2: 값 유효성 (-20 ~ 60) - 수집 구간만
        if len(df_target) > 0:
            valid_data = df_target[df_target[col].notna()]
            invalid_mask = (valid_data[col] < -20.0) | (valid_data[col] > 60.0)
            invalid_count = invalid_mask.sum()
            invalid_indices = valid_data[invalid_mask].index.tolist()
        else:
            invalid_count = 0
            invalid_indices = []
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('19-2', col, '값 유효성', '-20.0 ≤ module_avg_temp ≤ 60.0', status, invalid_count,
                       f'{invalid_count}건 범위 초과{window_note}' if invalid_count > 0 else f'정상{window_note}',
                       invalid_indices)
    
    def _validate_20_battery_modules(self):
        """20. battery_module_N_temperature 검증 (vehicle_master 기반 개수 확인)
        
        수집 구간: IGN 1→0 ~sleep, 충전중(1|0/0|1), 충전종료→0|0 ~sleep
        """
        # battery_module로 시작하고 temperature로 끝나는 컬럼 찾기
        module_cols = [col for col in self.df.columns 
                      if col.startswith('battery_module_') and col.endswith('_temperature')]
        
        expected_count = self.module_count
        criteria_count = f'{expected_count}개'
        
        if len(module_cols) == 0:
            self._add_result('20-1', 'battery_module_N_temperature', '수집 주기', '수집 구간 내 ≤20', 'N/A', 0, '컬럼 없음')
            self._add_result('20-2', 'battery_module_N_temperature', '데이터 개수', criteria_count, 'N/A', 0, '컬럼 없음')
            self._add_result('20-3', 'battery_module_N_temperature', '값 유효성(최소)', '-20.0 ≤ 개별 모듈 온도 최소값 ≤ 60.0', 'N/A', 0, '컬럼 없음')
            self._add_result('20-4', 'battery_module_N_temperature', '값 유효성(최대)', '-20.0 ≤ 개별 모듈 온도 최대값 ≤ 60.0', 'N/A', 0, '컬럼 없음')
            return
        
        # 수집 구간 마스크 (IGN OFF~sleep, 충전중, 충전종료~sleep)
        collection_mask = self._get_collection_window_mask()
        df_target = self.df[collection_mask]
        window_note = ' (수집구간: IGN OFF→sleep/충전중/충전종료→sleep)'
        
        # 20-1: 수집 주기 (수집 구간 내에서만)
        first_col = module_cols[0]
        fail_count, fail_indices = self._get_interval_violations_in_windows(first_col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('20-1', 'battery_module_N_temperature', '수집 주기', '수집 구간 내 ≤20',
                       status, fail_count,
                       f'{fail_count}건 위반{window_note}' if fail_count > 0 else f'정상{window_note}',
                       fail_indices)
        
        # 20-2: 데이터 개수 (vehicle_master 기반) - 실제 데이터가 있는 컬럼만 카운트
        # 전체 NaN인 컬럼은 제외 (스키마상 존재하지만 해당 차종에서 미사용)
        active_module_cols = [c for c in module_cols if self.df[c].notna().any()]
        actual_count = len(active_module_cols)
        if actual_count == expected_count:
            status = 'PASS'
            details = f'일치 (기준: {expected_count}개, 실제: {actual_count}개)'
            fail_cnt = 0
        elif actual_count > expected_count:
            status = 'PASS'
            details = f'기준 초과 (기준: {expected_count}개, 실제: {actual_count}개)'
            fail_cnt = 0
        else:
            status = 'FAIL'
            details = f'불일치 (기준: {expected_count}개, 실제: {actual_count}개)'
            fail_cnt = expected_count - actual_count
        self._add_result('20-2', 'battery_module_N_temperature', '데이터 개수', criteria_count,
                       status, fail_cnt, details)
        
        # 20-3: 최소값 유효성 (-20 ~ 60) - 수집 구간만
        if len(df_target) > 0:
            all_min_values = df_target[module_cols].min(axis=1)
            valid_data = all_min_values[all_min_values.notna()]
            invalid_mask = (valid_data < -20.0) | (valid_data > 60.0)
            invalid_count = invalid_mask.sum()
            invalid_indices = valid_data[invalid_mask].index.tolist()
        else:
            invalid_count = 0
            invalid_indices = []
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('20-3', 'battery_module_N_temperature', '값 유효성(최소)', '-20.0 ≤ 개별 모듈 온도 최소값 ≤ 60.0',
                       status, invalid_count,
                       f'{invalid_count}건 범위 초과{window_note}' if invalid_count > 0 else f'정상{window_note}',
                       invalid_indices)
        
        # 20-4: 최대값 유효성 (-20 ~ 60) - 수집 구간만
        if len(df_target) > 0:
            all_max_values = df_target[module_cols].max(axis=1)
            valid_data = all_max_values[all_max_values.notna()]
            invalid_mask = (valid_data < -20.0) | (valid_data > 60.0)
            invalid_count = invalid_mask.sum()
            invalid_indices = valid_data[invalid_mask].index.tolist()
        else:
            invalid_count = 0
            invalid_indices = []
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('20-4', 'battery_module_N_temperature', '값 유효성(최대)', '-20.0 ≤ 개별 모듈 온도 최대값 ≤ 60.0',
                       status, invalid_count,
                       f'{invalid_count}건 범위 초과{window_note}' if invalid_count > 0 else f'정상{window_note}',
                       invalid_indices)
    
    def _validate_21_oper_second(self):
        """21. oper_second 검증"""
        col = 'oper_second'
        
        if not self._check_column_exists(col):
            self._add_result('21-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            return
        
        # 21-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('21-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
    
    def _validate_22_acc_chg_ah(self):
        """22. acc_chg_ah 검증 (Ray 차종은 미수집 → N/A)"""
        col = 'acc_chg_ah'
        
        # Ray 차종은 누적 충/방전량 미수집 → N/A 처리
        if self.model_name == 'RAY':
            self._add_result('22-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '차종 미수집 (RAY)')
            self._add_result('22-2', col, '값 유효성', '0 ≤ acc_chg_ah ≤ 1,000,000', 'N/A', 0, '차종 미수집 (RAY)')
            self._add_result('22-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1', 'N/A', 0, '차종 미수집 (RAY)')
            return
        
        if not self._check_column_exists(col):
            self._add_result('22-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('22-2', col, '값 유효성', '0 ≤ acc_chg_ah ≤ 1,000,000', 'N/A', 0, '컬럼 없음')
            self._add_result('22-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1', 'N/A', 0, '컬럼 없음')
            return
        
        # 22-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('22-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 22-2: 값 유효성 (0 ~ 1,000,000)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < 0) | (valid_data[col] > 1000000)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('22-2', col, '값 유효성', '0 ≤ acc_chg_ah ≤ 1,000,000', status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
        
        # 22-3: 값 선형 증가
        valid_values = self.df[self.df[col].notna()][col].values
        if len(valid_values) >= 2:
            diffs = np.diff(valid_values)
            violations = np.where(diffs > 1)[0]
            fail_indices = [i + 1 for i in violations]
            status = 'PASS' if len(violations) == 0 else 'FAIL'
            self._add_result('22-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1',
                           status, len(violations),
                           f'{len(violations)}건 위반' if len(violations) > 0 else '정상',
                           fail_indices)
        else:
            self._add_result('22-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1',
                           'WARNING', 0, '데이터 부족')
    
    def _validate_23_acc_dchg_ah(self):
        """23. acc_dchg_ah 검증 (Fleet 기반 부호 반전, Ray 미수집)"""
        col = 'acc_dchg_ah'
        
        # Ray 차종은 누적 충/방전량 미수집 → N/A 처리
        if self.model_name == 'RAY':
            self._add_result('23-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '차종 미수집 (RAY)')
            self._add_result('23-2', col, '값 유효성', '0 ≤ acc_dchg_ah ≤ 1,000,000', 'N/A', 0, '차종 미수집 (RAY)')
            self._add_result('23-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1', 'N/A', 0, '차종 미수집 (RAY)')
            return
        
        # Fleet 기반 범위 설정 (VWGKALRT는 음수)
        if self.fleet == 'VWGKALRT':
            min_val, max_val = -1000000, 0
            criteria_text = '-1,000,000 ≤ acc_dchg_ah ≤ 0'
        else:
            min_val, max_val = 0, 1000000
            criteria_text = '0 ≤ acc_dchg_ah ≤ 1,000,000'
        
        if not self._check_column_exists(col):
            self._add_result('23-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('23-2', col, '값 유효성', criteria_text, 'N/A', 0, '컬럼 없음')
            self._add_result('23-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1', 'N/A', 0, '컬럼 없음')
            return
        
        # 23-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('23-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 23-2: 값 유효성 (Fleet 기반)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < min_val) | (valid_data[col] > max_val)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('23-2', col, '값 유효성', criteria_text, status, invalid_count,
                       f'{invalid_count}건 범위 초과 (Fleet: {self.fleet})' if invalid_count > 0 else '정상',
                       invalid_indices)
        
        # 23-3: 값 선형 증가 (Fleet에 따라 증가 방향 반전)
        valid_values = self.df[self.df[col].notna()][col].values
        if len(valid_values) >= 2:
            diffs = np.diff(valid_values)
            # VWGKALRT: 감소 방향 (-1 이상), 일반: 증가 방향 (1 이하)
            if self.fleet == 'VWGKALRT':
                violations = np.where(diffs < -1)[0]  # 감소폭이 1 초과
            else:
                violations = np.where(diffs > 1)[0]   # 증가폭이 1 초과
            fail_indices = [i + 1 for i in violations]
            status = 'PASS' if len(violations) == 0 else 'FAIL'
            diff_desc = '|현재 값 - NULL이 아닌 이전 값| ≤ 1'
            self._add_result('23-3', col, '값 선형 증가', diff_desc,
                           status, len(violations),
                           f'{len(violations)}건 위반 (Fleet: {self.fleet})' if len(violations) > 0 else '정상',
                           fail_indices)
        else:
            self._add_result('23-3', col, '값 선형 증가', '|현재 값 - NULL이 아닌 이전 값| ≤ 1',
                           'WARNING', 0, '데이터 부족')
    
    def _validate_24_acc_chg_wh(self):
        """24. acc_chg_wh 검증 (Ray 미수집)"""
        col = 'acc_chg_wh'
        
        # Ray 차종은 누적 충/방전량 미수집 → N/A 처리
        if self.model_name == 'RAY':
            self._add_result('24-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '차종 미수집 (RAY)')
            self._add_result('24-2', col, '값 유효성', '0 ≤ acc_chg_wh ≤ 1,000,000', 'N/A', 0, '차종 미수집 (RAY)')
            self._add_result('24-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1', 'N/A', 0, '차종 미수집 (RAY)')
            return
        
        if not self._check_column_exists(col):
            self._add_result('24-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('24-2', col, '값 유효성', '0 ≤ acc_chg_wh ≤ 1,000,000', 'N/A', 0, '컬럼 없음')
            self._add_result('24-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1', 'N/A', 0, '컬럼 없음')
            return
        
        # 24-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('24-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 24-2: 값 유효성
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < 0) | (valid_data[col] > 1000000)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('24-2', col, '값 유효성', '0 ≤ acc_chg_wh ≤ 1,000,000', status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
        
        # 24-3: 값 선형 증가
        valid_values = self.df[self.df[col].notna()][col].values
        if len(valid_values) >= 2:
            diffs = np.diff(valid_values)
            violations = np.where(diffs > 1)[0]
            fail_indices = [i + 1 for i in violations]
            status = 'PASS' if len(violations) == 0 else 'FAIL'
            self._add_result('24-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1',
                           status, len(violations),
                           f'{len(violations)}건 위반' if len(violations) > 0 else '정상',
                           fail_indices)
        else:
            self._add_result('24-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1',
                           'WARNING', 0, '데이터 부족')
    
    def _validate_25_acc_dchg_wh(self):
        """25. acc_dchg_wh 검증 (Fleet 기반 부호 반전, Ray 미수집)"""
        col = 'acc_dchg_wh'
        
        # Ray 차종은 누적 충/방전량 미수집 → N/A 처리
        if self.model_name == 'RAY':
            self._add_result('25-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '차종 미수집 (RAY)')
            self._add_result('25-2', col, '값 유효성', '0 ≤ acc_dchg_wh ≤ 1,000,000', 'N/A', 0, '차종 미수집 (RAY)')
            self._add_result('25-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1', 'N/A', 0, '차종 미수집 (RAY)')
            return
        
        # Fleet 기반 범위 설정 (VWGKALRT는 음수)
        if self.fleet == 'VWGKALRT':
            min_val, max_val = -1000000, 0
            criteria_text = '-1,000,000 ≤ acc_dchg_wh ≤ 0'
        else:
            min_val, max_val = 0, 1000000
            criteria_text = '0 ≤ acc_dchg_wh ≤ 1,000,000'
        
        if not self._check_column_exists(col):
            self._add_result('25-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('25-2', col, '값 유효성', criteria_text, 'N/A', 0, '컬럼 없음')
            self._add_result('25-3', col, '값 선형 증가', '현재 값 - NULL이 아닌 이전 값 ≤ 1', 'N/A', 0, '컬럼 없음')
            return
        
        # 25-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('25-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 25-2: 값 유효성 (Fleet 기반)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < min_val) | (valid_data[col] > max_val)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('25-2', col, '값 유효성', criteria_text, status, invalid_count,
                       f'{invalid_count}건 범위 초과 (Fleet: {self.fleet})' if invalid_count > 0 else '정상',
                       invalid_indices)
        
        # 25-3: 값 선형 증가 (Fleet에 따라 증가 방향 반전)
        valid_values = self.df[self.df[col].notna()][col].values
        if len(valid_values) >= 2:
            diffs = np.diff(valid_values)
            # VWGKALRT: 감소 방향 (-1 이상), 일반: 증가 방향 (1 이하)
            if self.fleet == 'VWGKALRT':
                violations = np.where(diffs < -1)[0]  # 감소폭이 1 초과
            else:
                violations = np.where(diffs > 1)[0]   # 증가폭이 1 초과
            fail_indices = [i + 1 for i in violations]
            status = 'PASS' if len(violations) == 0 else 'FAIL'
            diff_desc = '|현재 값 - NULL이 아닌 이전 값| ≤ 1'
            self._add_result('25-3', col, '값 선형 증가', diff_desc,
                           status, len(violations),
                           f'{len(violations)}건 위반 (Fleet: {self.fleet})' if len(violations) > 0 else '정상',
                           fail_indices)
        else:
            self._add_result('25-3', col, '값 선형 증가', '|현재 값 - NULL이 아닌 이전 값| ≤ 1',
                           'WARNING', 0, '데이터 부족')
    
    def _validate_26_pack_pwr(self):
        """26. pack_pwr 검증 (pack_pwr = pack_curr × pack_volt 기반 동적 계산)"""
        col = 'pack_pwr'
        
        # pack_pwr 범위 = pack_curr 범위 × pack_volt 범위
        # pack_curr: -360.0 ~ 750.0
        # pack_volt: cell_volt(2.6~4.5) × cell_count
        cell_volt_min = 2.6
        cell_volt_max = 4.5
        pack_volt_min = cell_volt_min * self.cell_count
        pack_volt_max = cell_volt_max * self.cell_count
        pack_curr_min = -360.0
        pack_curr_max = 750.0
        
        # 최소/최대 조합 중 실제 최소·최대 산출
        candidates = [
            pack_curr_min * pack_volt_min,
            pack_curr_min * pack_volt_max,
            pack_curr_max * pack_volt_min,
            pack_curr_max * pack_volt_max,
        ]
        pwr_min = min(candidates)
        pwr_max = max(candidates)
        
        criteria_text = (f'{pwr_min:,.1f} ≤ pack_pwr ≤ {pwr_max:,.1f} '
                        f'(curr {pack_curr_min}~{pack_curr_max}A × volt {pack_volt_min:.1f}~{pack_volt_max:.1f}V)')
        
        if not self._check_column_exists(col):
            self._add_result('26-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('26-2', col, '값 유효성', criteria_text, 'N/A', 0, '컬럼 없음')
            return
        
        # 26-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('26-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 26-2: 값 유효성 (동적 범위)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < pwr_min) | (valid_data[col] > pwr_max)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('26-2', col, '값 유효성', criteria_text, status, invalid_count,
                       f'{invalid_count}건 범위 초과 (셀 개수: {self.cell_count})' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_27_ir(self):
        """27. ir 검증"""
        col = 'ir'
        
        if not self._check_column_exists(col):
            self._add_result('27-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            return
        
        # 27-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('27-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
    
    def _validate_28_charging_time(self):
        """28. 충전 전류 소요시간 검증
        
        충전 시작 후 첫 충전 부호 발견 시간: -5초 ≤ time ≤ 25초
        - VWGKALRT: pack_curr > 1 (방전이 양수)
        - 기타 Fleet: pack_curr < -1 (충전이 음수)
        """
        required_cols = ['chg_conr_status_list', 'pack_curr', 'unix_time']
        
        criteria_text = '-5초 ≤ 충전 시작 후 첫 충전부호 발견 시간 ≤ 25초'
        
        if not all(self._check_column_exists(col) for col in required_cols):
            self._add_result('28-1', '충전 전류 소요시간', '값 유효성', 
                           criteria_text, 'N/A', 0, '필요한 컬럼 없음')
            return
        
        df = self.df
        chg_values = df['chg_conr_status_list'].astype(str)
        charging_mask = chg_values.isin(['1|0', '0|1'])
        
        if not charging_mask.any():
            self._add_result('28-1', '충전 전류 소요시간', '값 유효성',
                           criteria_text, 'N/A', 0, '충전 데이터 없음')
            return
        
        # 충전 구간 시작 시점 찾기
        charging_starts = df[charging_mask & ~charging_mask.shift(1, fill_value=False)].index
        fail_count = 0
        fail_indices = []
        
        for start_idx in charging_starts:
            search_end = min(start_idx + 100, len(df))
            search_df = df.loc[start_idx:search_end]
            
            # Fleet에 따라 충전 부호 방향 분기
            if self.fleet == 'VWGKALRT':
                curr_match = search_df[search_df['pack_curr'] > 1.0]
            else:
                curr_match = search_df[search_df['pack_curr'] < -1.0]
            
            if len(curr_match) > 0:
                first_match_idx = curr_match.index[0]
                time_diff = df.loc[first_match_idx, 'unix_time'] - df.loc[start_idx, 'unix_time']
                
                # -5초 ~ 25초 범위 밖이면 FAIL
                if time_diff < -5 or time_diff > 25:
                    fail_count += 1
                    fail_indices.append(start_idx)
            # 충전 부호가 아예 발견되지 않은 경우는 PASS (전류 변경 전 구간)
        
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('28-1', '충전 전류 소요시간', '값 유효성',
                       criteria_text,
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
    
    def _split_into_events(self):
        """unix_time 간격 ≥ 60초 기준으로 데이터를 이벤트(key cycle)별로 분할"""
        if 'unix_time' not in self.df.columns:
            return [self.df]
        
        df = self.df
        valid_mask = df['unix_time'].notna()
        if valid_mask.sum() < 2:
            return [df]
        
        times = df['unix_time'].values
        gaps = np.diff(times)
        # 60초 이상 간격이면 다른 key cycle
        split_points = np.where(gaps >= 60)[0] + 1
        
        if len(split_points) == 0:
            return [df]
        
        events = []
        prev = 0
        for sp in split_points:
            events.append(df.iloc[prev:sp].copy())
            prev = sp
        events.append(df.iloc[prev:].copy())
        
        return events
    
    def _validate_29_start_end_status(self):
        """29. 이벤트별 시작/종료 status 검증"""
        required_cols = ['ignit_status', 'chg_conr_status_list', 'main_relay_status']
        
        if not all(self._check_column_exists(col) for col in required_cols):
            self._add_result('29-1', '데이터 시작/종료 status', '값 유효성',
                           '이벤트별 시작/종료 상태 조합 유효성',
                           'N/A', 0, '필요한 컬럼 없음')
            return
        
        events = self._split_into_events()
        valid_start_tuples = [(0,0,0), (0,1,1), (1,0,0), (1,0,1), (0,0,1), (0,2,1), (1,1,1)]
        valid_end_tuples = [(0,0), (0,1), (0,2)]
        
        fail_count = 0
        fail_details = []
        
        for evt_idx, evt_df in enumerate(events, 1):
            valid_rows = evt_df[evt_df['ignit_status'].notna()]
            if len(valid_rows) == 0:
                continue
            
            first_idx = valid_rows.index[0]
            last_idx = valid_rows.index[-1]

            # 시작 상태 — relay/chg NaN이면 이 이벤트 스킵
            ign_start   = evt_df.loc[first_idx, 'ignit_status']
            relay_start = evt_df.loc[first_idx, 'main_relay_status']
            chg_start   = str(evt_df.loc[first_idx, 'chg_conr_status_list'])
            if pd.isna(relay_start) or pd.isna(ign_start):
                continue
            try:
                chg_val_start = int(chg_start.split('|')[0]) if '|' in chg_start else int(float(chg_start))
            except (ValueError, TypeError):
                continue
            start_tuple = (int(ign_start), chg_val_start, int(relay_start))

            # 종료 상태 — NaN이면 이 이벤트 스킵
            ign_end = evt_df.loc[last_idx, 'ignit_status']
            chg_end = str(evt_df.loc[last_idx, 'chg_conr_status_list'])
            if pd.isna(ign_end):
                continue
            try:
                chg_val_end = int(chg_end.split('|')[0]) if '|' in chg_end else int(float(chg_end))
            except (ValueError, TypeError):
                continue
            end_tuple = (int(ign_end), chg_val_end)
            
            start_valid = start_tuple in valid_start_tuples
            end_valid = end_tuple in valid_end_tuples
            
            if not start_valid or not end_valid:
                fail_count += 1
                msgs = []
                if not start_valid:
                    msgs.append(f'시작{start_tuple}')
                if not end_valid:
                    msgs.append(f'종료{end_tuple}')
                fail_details.append(f'이벤트{evt_idx}(row {first_idx}~{last_idx}): {", ".join(msgs)}')
        
        if fail_count == 0:
            status = 'PASS'
            details = f'정상 (총 {len(events)}개 이벤트)'
        else:
            status = 'FAIL'
            details = f'{fail_count}/{len(events)}개 이벤트 위반 | ' + ' | '.join(fail_details)
        
        self._add_result('29-1', '데이터 시작/종료 status', '값 유효성',
                       '이벤트별 시작/종료 상태 조합 유효성',
                       status, fail_count, details)
    
    def _validate_30_sleep_latency(self):
        """30. OBD Sleep Latency 검증 - 모든 이벤트(key cycle)별 검증
        
        자정 넘김 예외: 이벤트 종료 시점에 IGN=1이거나 충전 중(1|0/0|1)인 경우
        자정으로 인해 key cycle이 분리된 것으로 판단하여 검증에서 제외.
        """
        required_cols = ['ignit_status', 'chg_conr_status_list', 'unix_time']
        
        if not all(self._check_column_exists(col) for col in required_cols):
            self._add_result('30-1', 'OBD Sleep Latency', '값 유효성',
                           '이벤트별 시동OFF/충전종료 ~ 데이터종단 ≤ 35분',
                           'N/A', 0, '필요한 컬럼 없음')
            return
        
        events = self._split_into_events()
        
        fail_count = 0
        event_results = []
        
        for evt_idx, evt_df in enumerate(events, 1):
            if len(evt_df) < 2:
                continue
            
            evt_df = evt_df.reset_index(drop=True)
            last_time = evt_df['unix_time'].iloc[-1]
            
            # --- 자정 넘김 예외 처리 ---
            # 이벤트 마지막 행이 IGN=1 또는 충전중이면 → 아직 종료 이벤트가
            # 발생하지 않은 채 자정 등으로 key cycle이 분리된 것이므로 skip
            last_ign = evt_df['ignit_status'].iloc[-1]
            last_chg = str(evt_df['chg_conr_status_list'].iloc[-1])
            if last_ign == 1 or last_chg in ('1|0', '0|1'):
                event_results.append(f'이벤트{evt_idx}: 종료미완(IGN={int(last_ign)},CHG={last_chg}) SKIP')
                continue
            
            # --- 이벤트 내 종료 시점 탐색 ---
            # 시동 OFF (1 -> 0)
            ign_off_times = []
            ign_changes = evt_df['ignit_status'].diff()
            ign_off_indices = evt_df[ign_changes == -1].index
            if len(ign_off_indices) > 0:
                ign_off_times = evt_df.loc[ign_off_indices, 'unix_time'].tolist()
            
            # 충전 종료 (1|0/0|1 -> 0|0)
            chg_end_times = []
            chg_values = evt_df['chg_conr_status_list'].astype(str)
            for i in range(1, len(evt_df)):
                prev_chg = str(chg_values.iloc[i-1])
                curr_chg = str(chg_values.iloc[i])
                if prev_chg in ('1|0', '0|1') and curr_chg == '0|0':
                    chg_end_times.append(evt_df.loc[evt_df.index[i], 'unix_time'])
            
            # 마지막 종료 시점 결정
            all_end_times = ign_off_times + chg_end_times
            
            if len(all_end_times) == 0:
                event_results.append(f'이벤트{evt_idx}: 종료이벤트 없음')
                continue
            
            last_end_time = max(all_end_times)
            latency_seconds = last_time - last_end_time
            latency_minutes = latency_seconds / 60
            
            if latency_minutes <= 35:
                event_results.append(f'이벤트{evt_idx}: {latency_minutes:.1f}분 OK')
            else:
                fail_count += 1
                event_results.append(f'이벤트{evt_idx}: {latency_minutes:.1f}분 FAIL')
        
        if len(event_results) == 0:
            self._add_result('30-1', 'OBD Sleep Latency', '값 유효성',
                           '이벤트별 시동OFF/충전종료 ~ 데이터종단 ≤ 35분',
                           'WARNING', 0, '종료 이벤트 없음')
            return
        
        overall_status = 'PASS' if fail_count == 0 else 'FAIL'
        details = f'총 {len(events)}개 이벤트 | ' + ' | '.join(event_results)
        
        self._add_result('30-1', 'OBD Sleep Latency', '값 유효성',
                       '이벤트별 시동OFF/충전종료 ~ 데이터종단 ≤ 35분',
                       overall_status, fail_count, details)
    
    # 31번 항목 삭제됨 (이전 post-op 여부) → soh_rate가 31번으로 재배치
    
    def _validate_31_soh_rate(self):
        """31. soh_rate 검증"""
        col = 'soh_rate'
        
        if not self._check_column_exists(col):
            self._add_result('31-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0', 'N/A', 0, '컬럼 없음')
            self._add_result('31-2', col, '값 유효성', '50 ≤ soh_rate ≤ 100.0', 'N/A', 0, '컬럼 없음')
            return
        
        # 31-1: 수집 주기
        fail_count, fail_indices = self._get_interval_violations(col, 20)
        status = 'PASS' if fail_count == 0 else 'FAIL'
        self._add_result('31-1', col, '수집 주기', '현재 인덱스 - Null이 아닌 이전 인덱스 ≤ 20.0',
                       status, fail_count,
                       f'{fail_count}건 위반' if fail_count > 0 else '정상',
                       fail_indices)
        
        # 31-2: 값 유효성 (50 ~ 100)
        valid_data = self.df[self.df[col].notna()]
        invalid_mask = (valid_data[col] < 50) | (valid_data[col] > 100.0)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('31-2', col, '값 유효성', '50 ≤ soh_rate ≤ 100.0', status, invalid_count,
                       f'{invalid_count}건 범위 초과' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_32_ign_speed(self):
        """32. IGN * vehicle speed 검증"""
        required_cols = ['ignit_status', 'em_speed_kmh']
        
        if not all(self._check_column_exists(col) for col in required_cols):
            self._add_result('32-1', 'IGN * vehicle speed', '값 유효성',
                           'vehicle speed > 1 일때 IGN <> 1',
                           'N/A', 0, '필요한 컬럼 없음')
            return
        
        df = self.df.copy()
        # vehicle speed 255는 0으로 치환 (오류값 예외처리)
        df['em_speed_kmh'] = df['em_speed_kmh'].replace(255, 0)
        
        # vehicle speed > 1인데 IGN <> 1 (IGN이 1이 아닌 경우 - 오류 조건)
        invalid_mask = (df['em_speed_kmh'] > 1) & (df['ignit_status'] != 1)
        invalid_count = invalid_mask.sum()
        invalid_indices = df[invalid_mask].index.tolist()
        
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('32-1', 'IGN * vehicle speed', '값 유효성',
                       'vehicle speed > 1 일때 IGN <> 1',
                       status, invalid_count,
                       f'{invalid_count}건 위반' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_33_ign_relay(self):
        """33. IGN * main relay 검증"""
        required_cols = ['ignit_status', 'main_relay_status']
        
        if not all(self._check_column_exists(col) for col in required_cols):
            self._add_result('33-1', 'IGN * main relay', '값 유효성',
                           'IGN = 1 일때 main relay <> 1',
                           'N/A', 0, '필요한 컬럼 없음')
            return
        
        df = self.df
        # IGN = 1인데 main relay != 1 → 오류 (전체 행 대상 검증)
        valid_data = df[df['ignit_status'].notna() & df['main_relay_status'].notna()]
        invalid_mask = (valid_data['ignit_status'] == 1) & (valid_data['main_relay_status'] != 1)
        invalid_count = invalid_mask.sum()
        invalid_indices = valid_data[invalid_mask].index.tolist()
        
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        details = f'{invalid_count}건 위반 (전체 {len(valid_data)}행 중)' if invalid_count > 0 else f'정상 (전체 {len(valid_data)}행 검증)'
        self._add_result('33-1', 'IGN * main relay', '값 유효성',
                       'IGN = 1 일때 main relay <> 1',
                       status, invalid_count,
                       details,
                       invalid_indices)
    
    def _validate_34_chg_relay(self):
        """34. chg_conr_status_list * main relay 검증
        
        34-1: chg_conr_status_list ≥ 1 일때 main_relay < 1 검출
        """
        required_cols = ['chg_conr_status_list', 'main_relay_status']
        
        if not all(self._check_column_exists(col) for col in required_cols):
            self._add_result('34-1', 'chg * main relay', '값 유효성',
                           '1 ≤ chg_conr_status_list 일때 main relay < 1',
                           'N/A', 0, '필요한 컬럼 없음')
            return
        
        df = self.df
        chg_str = df['chg_conr_status_list'].astype(str)
        chg_is_charging = chg_str.isin(['1|0', '0|1'])
        
        # chg가 '1|0' 또는 '0|1'이면서 main_relay_status < 1 (즉 0)인 경우
        invalid_mask = chg_is_charging & (df['main_relay_status'] < 1)
        invalid_count = invalid_mask.sum()
        invalid_indices = df[invalid_mask].index.tolist()
        
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('34-1', 'chg * main relay', '값 유효성',
                       '1 ≤ chg_conr_status_list 일때 main relay < 1',
                       status, invalid_count,
                       f'{invalid_count}건 위반' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    def _validate_35_chg_condition(self):
        """35. chg_conr_status_list 충전 조건 검증
        
        35-1: speed=0, relay=1, 전류 ≥ 1 일때 chg_conr_status_list가 충전 상태인지
          - VWGKALRT: pack_curr > 1 (방전이 양수)
          - 기타 Fleet: pack_curr < -1 (충전이 음수)
        """
        required_cols = ['em_speed_kmh', 'main_relay_status', 'pack_curr', 'chg_conr_status_list']
        
        if not all(self._check_column_exists(col) for col in required_cols):
            self._add_result('35-1', 'chg_conr_status_list', '값 유효성',
                           'speed=0, relay=1, 전류조건 → chg≥1',
                           'N/A', 0, '필요한 컬럼 없음')
            return
        
        df = self.df
        chg_str = df['chg_conr_status_list'].astype(str)
        chg_is_charging = chg_str.isin(['1|0', '0|1'])
        
        # Fleet에 따라 전류 조건 분기
        if self.fleet == 'VWGKALRT':
            condition_mask = (
                (df['em_speed_kmh'] == 0) &
                (df['main_relay_status'] == 1) &
                (df['pack_curr'] > 1)
            )
            criteria_text = 'speed=0, relay=1, curr>1 → chg≥1 (VWGKALRT)'
        else:
            condition_mask = (
                (df['em_speed_kmh'] == 0) &
                (df['main_relay_status'] == 1) &
                (df['pack_curr'] < -1)
            )
            criteria_text = 'speed=0, relay=1, curr<-1 → chg≥1'
        
        # 조건 만족하는데 chg가 충전 상태가 아닌 경우
        invalid_mask = condition_mask & ~chg_is_charging
        invalid_count = invalid_mask.sum()
        invalid_indices = df[invalid_mask].index.tolist()
        
        status = 'PASS' if invalid_count == 0 else 'FAIL'
        self._add_result('35-1', 'chg_conr_status_list', '값 유효성',
                       criteria_text,
                       status, invalid_count,
                       f'{invalid_count}건 위반' if invalid_count > 0 else '정상',
                       invalid_indices)
    
    # ==================== 리포트 생성 ====================
    
    def generate_report(self, output_path, source_file_path=None):
        """
        검증 결과를 XLSX 파일로 저장 (서식 적용)
        :param output_path: 출력 파일 경로
        :param source_file_path: 원본 데이터 파일 경로 (하이퍼링크용)
        """
        if not self.results:
            logger.warning("저장할 검증 결과가 없습니다.")
            return

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        ws = wb.active
        ws.title = "BMS Validation"

        # DataFrame 생성
        report_df = pd.DataFrame(self.results)

        # DataFrame을 시트에 쓰기
        for r_idx, row in enumerate(dataframe_to_rows(report_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

        # 테두리 스타일
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 1. 전체 테두리 적용
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 2. 헤더(1행) 주황색 배경
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # 3. ID 컬럼 병합 (Column이 같은 행끼리)
        merge_ranges = []
        current_column_value = None
        start_row = 2

        for row_idx in range(2, ws.max_row + 2):
            if row_idx <= ws.max_row:
                column_value = ws.cell(row=row_idx, column=2).value
            else:
                column_value = None

            if column_value != current_column_value:
                if current_column_value is not None and row_idx - 1 >= start_row:
                    if row_idx - 1 > start_row:
                        merge_ranges.append((start_row, row_idx - 1))

                current_column_value = column_value
                start_row = row_idx

        # ID 셀 병합 적용
        for start, end in merge_ranges:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            ws.cell(row=start, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # 4. Status 컬럼 색상 적용 (5번째 컬럼)
        status_col_idx = 5

        for row_idx in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row_idx, column=status_col_idx)
            status_value = status_cell.value

            if status_value == 'PASS':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100")
            elif status_value == 'FAIL':
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(color="9C0006", bold=True)
            elif status_value == 'WARNING':
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                status_cell.font = Font(color="9C6500")

        # 열 너비 자동 조정
        column_widths = {
            1: 10,   # ID
            2: 25,   # Column
            3: 30,   # Check
            4: 50,   # Criteria
            5: 12,   # Status
            6: 12,   # Fail_Count
            7: 40    # Details
        }

        for col_idx, width in column_widths.items():
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        # 행 높이 설정
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20

        # 파일 저장
        wb.save(output_path)
        logger.info(f"검증 리포트 저장 완료: {output_path}")
        print(f"검증 리포트 저장: {output_path}")

    def generate_html_report(self, output_path, source_file_path=None):
        """HTML 리포트 생성 — bms_report_generator.py 위임 (신규 추가)"""
        from bms_report_generator import generate_html_report as _gen
        _gen(self.results, self.df, self.vehicle_model, self.fleet, output_path, source_file_path)
