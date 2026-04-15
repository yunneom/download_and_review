"""
bms_report_generator.py
BMS 검증 리포트 생성 모듈
- XLSX 리포트 (generate_excel_report)
- HTML 리포트  (generate_html_report)

BMSDataValidator로부터 분리된 양식 전용 모듈.
직접 호출하거나 BMSDataValidator.generate_report() / .generate_html_report() 래퍼를 통해 사용.
"""

import os
from logger import logger


def generate_excel_report(results, output_path):
    """
    검증 결과를 XLSX 파일로 저장 (서식 적용)
    :param results: BMSDataValidator.results (list of dict)
    :param output_path: 출력 파일 경로
    """
    if not results:
        logger.warning("저장할 검증 결과가 없습니다.")
        return

    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "BMS Validation"

    report_df = pd.DataFrame(results)

    for r_idx, row in enumerate(dataframe_to_rows(report_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # ID 컬럼 병합 (같은 Column 값끼리)
    merge_ranges = []
    current_column_value = None
    start_row = 2

    for row_idx in range(2, ws.max_row + 2):
        if row_idx <= ws.max_row:
            column_value = ws.cell(row=row_idx, column=2).value
        else:
            column_value = None

        if column_value != current_column_value:
            if current_column_value is not None and row_idx - 1 >= start_row:
                if row_idx - 1 > start_row:
                    merge_ranges.append((start_row, row_idx - 1))
            current_column_value = column_value
            start_row = row_idx

    for start, end in merge_ranges:
        ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        ws.cell(row=start, column=1).alignment = Alignment(horizontal='center', vertical='center')

    status_col_idx = 5
    for row_idx in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row_idx, column=status_col_idx)
        status_value = status_cell.value

        if status_value == 'PASS':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100")
        elif status_value == 'FAIL':
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006", bold=True)
        elif status_value == 'WARNING':
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            status_cell.font = Font(color="9C6500")

    column_widths = {
        1: 10,   # ID
        2: 25,   # Column
        3: 30,   # Check
        4: 50,   # Criteria
        5: 12,   # Status
        6: 12,   # Fail_Count
        7: 40    # Details
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20

    wb.save(output_path)
    logger.info(f"검증 리포트 저장 완료: {output_path}")
    print(f"검증 리포트 저장: {output_path}")


def generate_html_report(results, df, vehicle_model, fleet, output_path, source_file_path=None):
    """
    검증 결과를 HTML 파일로 저장 (브라우저에서 바로 열람 가능)
    :param results:          BMSDataValidator.results (list of dict)
    :param df:               원본 DataFrame (그래프용, None 가능)
    :param vehicle_model:    차량 모델명 문자열
    :param fleet:            Fleet 문자열
    :param output_path:      출력 HTML 파일 경로
    :param source_file_path: 원본 데이터 파일 경로 (헤더 표시용)
    """
    if not results:
        logger.warning("저장할 검증 결과가 없습니다.")
        return

    import pandas as pd
    from datetime import datetime

    counts = {'PASS': 0, 'FAIL': 0, 'WARNING': 0, 'N/A': 0}
    for r in results:
        s = r.get('Status', '')
        if s in counts:
            counts[s] += 1

    total = sum(counts.values())
    pass_pct = round(counts['PASS'] / total * 100) if total else 0

    vehicle_info = f"{vehicle_model or ''} {fleet or ''}".strip()
    source_name = os.path.basename(source_file_path) if source_file_path else ''
    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    STATUS_COLOR = {
        'PASS':    ('#00c4a0', 'rgba(0,196,160,0.12)'),
        'FAIL':    ('#ff5a5a', 'rgba(255,90,90,0.12)'),
        'WARNING': ('#f5a623', 'rgba(245,166,35,0.12)'),
        'N/A':     ('#7a8fa8', 'rgba(122,143,168,0.12)'),
    }

    def row_html(r):
        sid = r.get('ID', '')
        col = r.get('Column', '')
        chk = r.get('Check', '')
        cri = r.get('Criteria', '').replace('<', '&lt;').replace('>', '&gt;')
        sts = r.get('Status', '')
        cnt = r.get('Fail_Count', 0)
        det = str(r.get('Details', '')).replace('<', '&lt;').replace('>', '&gt;')
        fg, bg = STATUS_COLOR.get(sts, ('#e8f0fe', 'transparent'))
        fail_cnt_html = (
            f'<span style="color:{fg}; font-weight:700;">{cnt}</span>'
            if sts in ('FAIL', 'WARNING') and cnt else
            '<span style="color:#4a6a8a;">0</span>'
        )
        badge = (
            f'<span style="display:inline-block; padding:2px 10px; border-radius:10px;'
            f' font-size:11px; font-weight:700; background:{bg}; color:{fg};">{sts}</span>'
        )
        detail_html = ''
        if det and det not in ('정상', '0'):
            detail_html = (
                f'<div style="font-size:11px; color:#8a9abb; margin-top:3px;'
                f' font-family:Consolas,monospace;">{det}</div>'
            )
        return (
            f'<tr>'
            f'<td style="color:#4a7aaa; font-weight:700; white-space:nowrap;">{sid}</td>'
            f'<td style="font-family:Consolas,monospace; font-size:12px;">{col}</td>'
            f'<td>{chk}</td>'
            f'<td style="font-size:12px; color:#8a9abb;">{cri}</td>'
            f'<td style="text-align:center;">{badge}</td>'
            f'<td style="text-align:center;">{fail_cnt_html}</td>'
            f'<td><div style="font-size:12px;">{det if not detail_html else ""}</div>{detail_html}</td>'
            f'</tr>'
        )

    rows_html = '\n'.join(row_html(r) for r in results)

    # donut SVG arcs
    circ = 2 * 3.14159 * 26
    pass_arc = round(circ * counts['PASS']    / total, 1) if total else 0
    fail_arc = round(circ * counts['FAIL']    / total, 1) if total else 0
    warn_arc = round(circ * counts['WARNING'] / total, 1) if total else 0
    na_arc   = round(circ * counts['N/A']     / total, 1) if total else 0
    off_fail = -pass_arc
    off_warn = -(pass_arc + fail_arc)
    off_na   = -(pass_arc + fail_arc + warn_arc)

    # ── 그래프 데이터 준비 ──
    cdn_links  = ''
    graph_css  = ''
    graph_html = ''
    chart_js   = ''

    if df is not None and len(df) > 0:
        import json as _json

        MAX_PTS = 1500
        step = max(1, len(df) // MAX_PTS)
        dfs  = df.iloc[::step].reset_index(drop=True)
        total_rows  = len(df)
        sample_rows = len(dfs)

        if 'unix_time' in dfs.columns:
            try:
                ts = pd.to_datetime(dfs['unix_time'], unit='s', utc=True).dt.tz_convert('Asia/Seoul')
                x_vals = ts.dt.strftime('%H:%M:%S').tolist()
            except Exception:
                x_vals = dfs['unix_time'].astype(str).tolist()
        else:
            x_vals = list(range(sample_rows))

        CC = ['#00c4a0', '#7ab3f5', '#f5a623', '#ff7a7a', '#c07aee', '#5aafff', '#ffd700']

        GROUPS_DEF = [
            ('SOC',                ['soc_rate', 'soc_display_rate'],                          '%'),
            ('온도 (Temperature)',  ['module_min_temp', 'module_max_temp', 'module_avg_temp'], '°C'),
            ('거리·속도',           ['mile_km', 'em_speed_kmh'],                              ''),
            ('전류·전압 (Pack)',     ['pack_curr', 'pack_volt'],                              ''),
            ('셀전압 (Cell Volt)',   ['cell_min_volt', 'cell_max_volt'],                       'V'),
            ('기타 (상태·저항)',     ['ignit_status', 'main_relay_status', 'ir'],              ''),
        ]

        groups = []
        for gname, cols, unit in GROUPS_DEF:
            avail = [c for c in cols if c in dfs.columns]
            if not avail:
                continue
            series = []
            for i, col in enumerate(avail):
                vals = []
                for v in dfs[col]:
                    try:
                        fv = float(v)
                        vals.append(None if pd.isna(fv) else round(fv, 4))
                    except (TypeError, ValueError):
                        vals.append(None)
                series.append({'name': col, 'data': vals, 'color': CC[i % len(CC)]})
            groups.append({'name': gname, 'unit': unit, 'series': series})

        if groups:
            all_data = _json.dumps({'x': x_vals, 'groups': groups}, ensure_ascii=False)

            cdn_links = (
                '<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script>\n'
                '<script src="https://cdn.jsdelivr.net/npm/hammerjs@2.0.8/hammer.min.js"></script>\n'
                '<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-zoom@2.0.1/dist/chartjs-plugin-zoom.min.js"></script>'
            )

            graph_css = (
                '.graph-section{border-bottom:1px solid rgba(255,255,255,0.06);}'
                '.graph-sec-hdr{list-style:none;display:flex;align-items:center;gap:10px;'
                'padding:14px 40px;cursor:pointer;font-size:14px;font-weight:700;'
                'background:#0d1b2a;user-select:none;}'
                '.graph-sec-hdr::-webkit-details-marker{display:none;}'
                '.graph-body{background:#0a1520;}'
                '.gpanel{border-top:1px solid rgba(255,255,255,0.05);}'
                '.gpanel-hdr{list-style:none;display:flex;align-items:center;gap:8px;'
                'padding:10px 40px;cursor:pointer;font-size:13px;font-weight:600;'
                'background:#0d1b2a;user-select:none;}'
                '.gpanel-hdr::-webkit-details-marker{display:none;}'
                '.gp-arrow{font-size:10px;color:#8a9abb;transition:transform 0.2s;display:inline-block;}'
                '.gpanel[open]>.gpanel-hdr .gp-arrow{transform:rotate(90deg);}'
                '.gp-name{color:#e8f0fe;}'
                '.gp-unit{font-size:11px;color:#8a9abb;}'
                '.gp-note{font-size:10px;color:#4a6a8a;margin-left:auto;}'
                '.gpanel-body{padding:8px 40px 16px;background:#0a1520;}'
            )

            panels_html = ''
            for gi, g in enumerate(groups):
                open_attr = ' open' if gi == 0 else ''
                panels_html += (
                    f'<details class="gpanel"{open_attr}>'
                    f'<summary class="gpanel-hdr">'
                    f'<span class="gp-arrow">&#9658;</span>'
                    f'<span class="gp-name">{g["name"]}</span>'
                    f'<span class="gp-unit">&nbsp;{g["unit"]}</span>'
                    f'<span class="gp-note">휠: 확대/축소 &middot; 드래그: 이동 &middot; 더블클릭: 초기화</span>'
                    f'</summary>'
                    f'<div class="gpanel-body"><canvas id="gchart-{gi}" height="140"></canvas></div>'
                    f'</details>'
                )

            graph_html = (
                '<details class="graph-section" open>'
                '<summary class="graph-sec-hdr">&#128202; 데이터 그래프 '
                f'<span style="font-size:11px;color:#8a9abb;font-weight:400;">'
                f'(샘플 {sample_rows:,}행 / 전체 {total_rows:,}행)</span></summary>'
                f'<div class="graph-body">{panels_html}</div>'
                '</details>'
            )

            chart_js = '''<script>
(function(){
  var GDATA = ''' + all_data + ''';
  if(!GDATA||!GDATA.groups) return;
  var base = {
    animation: false,
    responsive: true,
    interaction:{mode:'index',intersect:false},
    plugins:{
      legend:{labels:{color:'#8a9abb',font:{size:11}}},
      tooltip:{backgroundColor:'#132236',titleColor:'#e8f0fe',bodyColor:'#8a9abb',
               callbacks:{label:function(ctx){return ' '+ctx.dataset.label+': '+ctx.parsed.y;}}},
      zoom:{pan:{enabled:true,mode:'x'},zoom:{wheel:{enabled:true},pinch:{enabled:true},mode:'x'}}
    },
    scales:{
      x:{ticks:{color:'#4a6a8a',maxTicksLimit:14,maxRotation:0},
         grid:{color:'rgba(255,255,255,0.04)'}},
      y:{ticks:{color:'#8a9abb'},grid:{color:'rgba(255,255,255,0.07)'}}
    }
  };
  GDATA.groups.forEach(function(g,gi){
    var canvas = document.getElementById('gchart-'+gi);
    if(!canvas) return;
    var datasets = g.series.map(function(s){
      return {label:s.name,data:s.data,borderColor:s.color,
              backgroundColor:s.color+'18',borderWidth:1.5,
              pointRadius:0,fill:false,tension:0.2,spanGaps:true};
    });
    var opts = JSON.parse(JSON.stringify(base));
    if(g.unit) opts.scales.y.title={display:true,text:g.unit,color:'#8a9abb'};
    var chart = new Chart(canvas,{type:'line',data:{labels:GDATA.x,datasets:datasets},options:opts});
    canvas.addEventListener('dblclick',function(){chart.resetZoom();});
  });
})();
</script>'''

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>BMS 검증 리포트 — {vehicle_info}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI','Malgun Gothic',sans-serif; background:#0d1b2a; color:#e8f0fe; }}
  .header {{ background: linear-gradient(135deg,#0f3055,#1a2e45); padding:28px 40px 22px; border-bottom:1px solid rgba(255,255,255,0.08); }}
  .header h1 {{ font-size:22px; font-weight:800; margin-bottom:6px; }}
  .header .meta {{ font-size:12px; color:#8a9abb; }}
  .summary {{ display:flex; gap:14px; padding:20px 40px; background:#132236; border-bottom:1px solid rgba(255,255,255,0.06); flex-wrap:wrap; align-items:center; }}
  .stat {{ text-align:center; padding:10px 18px; border-radius:10px; min-width:80px; }}
  .stat .num {{ font-size:28px; font-weight:800; line-height:1; }}
  .stat .lbl {{ font-size:11px; margin-top:4px; }}
  .donut {{ position:relative; width:72px; height:72px; margin-right:8px; flex-shrink:0; }}
  .donut svg {{ transform:rotate(-90deg); }}
  .donut .pct {{ position:absolute; inset:0; display:flex; align-items:center; justify-content:center; font-size:14px; font-weight:800; color:#00c4a0; }}
  .filter-bar {{ padding:14px 40px; display:flex; gap:8px; flex-wrap:wrap; align-items:center; background:#0d1b2a; border-bottom:1px solid rgba(255,255,255,0.05); }}
  .filter-bar span {{ font-size:12px; color:#8a9abb; margin-right:4px; }}
  .fbtn {{ padding:4px 14px; border-radius:16px; border:1px solid; background:transparent; cursor:pointer; font-size:12px; font-weight:600; transition:all 0.15s; }}
  .fbtn.active-pass {{ background:rgba(0,196,160,0.15); border-color:#00c4a0; color:#00c4a0; }}
  .fbtn.active-fail {{ background:rgba(255,90,90,0.15); border-color:#ff5a5a; color:#ff5a5a; }}
  .fbtn.active-warn {{ background:rgba(245,166,35,0.15); border-color:#f5a623; color:#f5a623; }}
  .fbtn.active-na   {{ background:rgba(122,143,168,0.15); border-color:#7a8fa8; color:#7a8fa8; }}
  .fbtn.off {{ border-color:rgba(255,255,255,0.15); color:rgba(255,255,255,0.25); }}
  .search {{ margin-left:auto; padding:5px 12px; border-radius:16px; border:1px solid rgba(255,255,255,0.12); background:rgba(255,255,255,0.04); color:#e8f0fe; font-size:12px; outline:none; }}
  .search::placeholder {{ color:#4a6a8a; }}
  .table-wrap {{ padding:20px 40px 60px; overflow-x:auto; }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  th {{ background:rgba(26,107,204,0.18); color:#00c4a0; font-weight:600; text-align:left; padding:10px 12px; font-size:11px; letter-spacing:0.5px; position:sticky; top:0; }}
  td {{ padding:9px 12px; border-bottom:1px solid rgba(255,255,255,0.04); vertical-align:top; }}
  tr:hover td {{ background:rgba(255,255,255,0.03); }}
  tr.hidden {{ display:none; }}
  .footer {{ text-align:center; padding:16px; font-size:11px; color:#4a6a8a; border-top:1px solid rgba(255,255,255,0.05); }}
  @media print {{
    .filter-bar, .footer, .graph-section {{ display:none; }}
    body {{ background:#fff; color:#000; }}
    table {{ font-size:11px; }}
    th {{ background:#dce6f0 !important; color:#000 !important; }}
  }}
  {graph_css}
</style>
{cdn_links}
</head>
<body>

<div class="header">
  <h1>🔋 BMS 데이터 검증 리포트</h1>
  <div class="meta">
    차량: <strong style="color:#e8f0fe;">{vehicle_info or 'N/A'}</strong>
    {('&nbsp;·&nbsp; 파일: <strong style="color:#e8f0fe;">' + source_name + '</strong>') if source_name else ''}
    &nbsp;·&nbsp; 생성: {generated_at}
  </div>
</div>

<div class="summary">
  <div class="donut">
    <svg viewBox="0 0 64 64" width="72" height="72">
      <circle cx="32" cy="32" r="26" fill="none" stroke="rgba(255,255,255,0.06)" stroke-width="10"/>
      <circle cx="32" cy="32" r="26" fill="none" stroke="#00c4a0" stroke-width="10"
              stroke-dasharray="{pass_arc} {round(circ - pass_arc,1)}" stroke-dashoffset="0"/>
      <circle cx="32" cy="32" r="26" fill="none" stroke="#ff5a5a" stroke-width="10"
              stroke-dasharray="{fail_arc} {round(circ - fail_arc,1)}" stroke-dashoffset="{off_fail}"/>
      <circle cx="32" cy="32" r="26" fill="none" stroke="#f5a623" stroke-width="10"
              stroke-dasharray="{warn_arc} {round(circ - warn_arc,1)}" stroke-dashoffset="{off_warn}"/>
      <circle cx="32" cy="32" r="26" fill="none" stroke="#7a8fa8" stroke-width="10"
              stroke-dasharray="{na_arc} {round(circ - na_arc,1)}" stroke-dashoffset="{off_na}"/>
    </svg>
    <div class="pct">{pass_pct}%</div>
  </div>
  <div class="stat" style="background:rgba(0,196,160,0.10);">
    <div class="num" style="color:#00c4a0;">{counts['PASS']}</div>
    <div class="lbl" style="color:#00c4a0;">PASS</div>
  </div>
  <div class="stat" style="background:rgba(255,90,90,0.10);">
    <div class="num" style="color:#ff5a5a;">{counts['FAIL']}</div>
    <div class="lbl" style="color:#ff5a5a;">FAIL</div>
  </div>
  <div class="stat" style="background:rgba(245,166,35,0.10);">
    <div class="num" style="color:#f5a623;">{counts['WARNING']}</div>
    <div class="lbl" style="color:#f5a623;">WARNING</div>
  </div>
  <div class="stat" style="background:rgba(122,143,168,0.10);">
    <div class="num" style="color:#7a8fa8;">{counts['N/A']}</div>
    <div class="lbl" style="color:#7a8fa8;">N/A</div>
  </div>
  <div style="margin-left:auto; font-size:12px; color:#8a9abb; text-align:right;">
    총 {total}개 항목<br>
    <span style="color:#00c4a0; font-weight:700;">{'검증 통과' if counts['FAIL'] == 0 else str(counts['FAIL']) + '개 항목 실패'}</span>
  </div>
</div>

{graph_html}

<div class="filter-bar">
  <span>필터:</span>
  <button class="fbtn active-pass" onclick="toggle(this,'PASS')">PASS</button>
  <button class="fbtn active-fail" onclick="toggle(this,'FAIL')">FAIL</button>
  <button class="fbtn active-warn" onclick="toggle(this,'WARNING')">WARNING</button>
  <button class="fbtn active-na"   onclick="toggle(this,'N/A')">N/A</button>
  <input class="search" type="text" placeholder="ID / Column 검색..." oninput="search(this.value)">
</div>

<div class="table-wrap">
  <table id="result-table">
    <thead>
      <tr>
        <th>ID</th>
        <th>Column</th>
        <th>Check</th>
        <th>Criteria</th>
        <th style="text-align:center;">Status</th>
        <th style="text-align:center;">Fail Count</th>
        <th>Details</th>
      </tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
  </table>
</div>

<div class="footer">
  DART · BMS Data Quality Assurance System &nbsp;·&nbsp; {generated_at}
</div>

<script>
  const active = {{PASS:true, FAIL:true, WARNING:true, 'N/A':true}};
  const classMap = {{PASS:'active-pass', FAIL:'active-fail', WARNING:'active-warn', 'N/A':'active-na'}};
  let searchTerm = '';

  function toggle(btn, status) {{
    active[status] = !active[status];
    btn.className = 'fbtn ' + (active[status] ? classMap[status] : 'off');
    applyFilter();
  }}

  function search(val) {{
    searchTerm = val.toLowerCase();
    applyFilter();
  }}

  function applyFilter() {{
    document.querySelectorAll('#result-table tbody tr').forEach(row => {{
      const cells = row.querySelectorAll('td');
      if (!cells.length) return;
      const id  = (cells[0]?.textContent || '').toLowerCase();
      const col = (cells[1]?.textContent || '').toLowerCase();
      const sts = (cells[4]?.textContent || '').trim();
      const matchFilter = active[sts] !== false;
      const matchSearch = !searchTerm || id.includes(searchTerm) || col.includes(searchTerm);
      row.classList.toggle('hidden', !(matchFilter && matchSearch));
    }});
  }}
</script>
{chart_js}
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    logger.info(f"HTML 리포트 저장 완료: {output_path}")
    print(f"HTML 리포트 저장: {output_path}")
