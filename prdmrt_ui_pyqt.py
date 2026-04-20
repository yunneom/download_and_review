"""
PRDMRT BMS 검증 도구 - GUI (PyQt5)
- S3에서 parquet/csv 파일 다운로드
- BMS 37항목 검증 실행 및 결과 표시
- 검증 결과 xlsx 보고서 생성
- PID/날짜 기반 배치 검증 지원
"""

import sys
import os
import pandas as pd
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTextEdit, QProgressBar,
    QMessageBox, QGroupBox, QRadioButton, QButtonGroup, QFileDialog,
    QMenuBar, QMenu, QAction, QCalendarWidget, QDialog, QDialogButtonBox,
    QPlainTextEdit, QScrollArea, QTableWidget, QHeaderView,
)
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QDate
from PyQt5.QtGui import QFont, QColor

# 모듈화된 핵심 클래스 import
from prdmrt_core import PrdmrtCore
from logger import logger
from data_validator_bms import BMSDataValidator
from config import AWS_CONFIG

# AWS 설정 (.env에서 로드)
AWS_ACCESS_KEY_ID = AWS_CONFIG['access_key']
AWS_SECRET_ACCESS_KEY = AWS_CONFIG['secret_key']
AWS_DEFAULT_REGION = AWS_CONFIG['region']


class DateRangeDialog(QDialog):
    """날짜 범위 선택 다이얼로그"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("기간 선택")
        self.setModal(True)
        
        layout = QVBoxLayout()
        
        # 시작일
        start_group = QGroupBox("시작일")
        start_layout = QVBoxLayout()
        self.start_calendar = QCalendarWidget()
        self.start_calendar.setGridVisible(True)
        start_layout.addWidget(self.start_calendar)
        start_group.setLayout(start_layout)
        layout.addWidget(start_group)
        
        # 종료일
        end_group = QGroupBox("종료일")
        end_layout = QVBoxLayout()
        self.end_calendar = QCalendarWidget()
        self.end_calendar.setGridVisible(True)
        end_layout.addWidget(self.end_calendar)
        end_group.setLayout(end_layout)
        layout.addWidget(end_group)
        
        # 버튼
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
        self.setLayout(layout)
    
    def get_date_range(self):
        """선택된 날짜 범위 반환"""
        start_date = self.start_calendar.selectedDate()
        end_date = self.end_calendar.selectedDate()
        return start_date.toString("yyyy-MM-dd"), end_date.toString("yyyy-MM-dd")


class DataValidationDialog(QDialog):
    """데이터 검증 다이얼로그"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("고급 데이터 검증")
        self.setModal(True)
        self.setMinimumSize(900, 700)
        
        self.validation_rules = {}
        self.csv_file = None
        
        layout = QVBoxLayout()
        
        # 파일 선택
        file_group = QGroupBox("1. 데이터 파일 선택 (CSV 또는 Parquet)")
        file_layout = QHBoxLayout()
        
        self.file_path_input = QLineEdit()
        self.file_path_input.setReadOnly(True)
        self.file_path_input.setPlaceholderText("CSV 또는 Parquet 파일을 선택하세요...")
        file_layout.addWidget(self.file_path_input)
        
        select_file_btn = QPushButton("파일 선택")
        select_file_btn.clicked.connect(self.select_csv_file)
        file_layout.addWidget(select_file_btn)
        
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)
        
        # 검증 규칙 설정
        rules_group = QGroupBox("2. 검증 규칙 설정 (컬럼명, 체크항목, 설명)")
        rules_layout = QVBoxLayout()
        
        # 규칙 추가 영역
        self.rules_text = QTextEdit()
        self.rules_text.setPlaceholderText(
            "검증 규칙 예시:\n\n"
            "ignit_status,valid_values=[0;1],점화 상태\n"
            "soc_display_rate,range=[0;100],SOC 값\n"
            "soh_rate,range=[0;100],SOH 값\n"
            "module_min_temp,range=[-40;85],최저 온도\n"
            "pack_volt,range=[200;500],팩 전압\n\n"
            "형식: 컬럼명,체크타입=값,설명\n"
            "  - valid_values=[값1;값2;...] : 허용값 지정\n"
            "  - range=[최소;최대] : 범위 지정\n"
            "  - null : NULL 체크만"
        )
        self.rules_text.setMinimumHeight(200)
        rules_layout.addWidget(self.rules_text)
        
        # 기본 템플릿 로드 버튼
        template_btn_layout = QHBoxLayout()
        load_template_btn = QPushButton("기본 템플릿 로드")
        load_template_btn.clicked.connect(self.load_default_template)
        template_btn_layout.addWidget(load_template_btn)
        template_btn_layout.addStretch()
        rules_layout.addLayout(template_btn_layout)
        
        rules_group.setLayout(rules_layout)
        layout.addWidget(rules_group)
        
        # 실행 버튼
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        validate_btn = QPushButton("검증 실행")
        validate_btn.setMinimumSize(120, 40)
        validate_btn.clicked.connect(self.run_validation)
        validate_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                font-size: 14px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        btn_layout.addWidget(validate_btn)
        
        close_btn = QPushButton("닫기")
        close_btn.setMinimumSize(120, 40)
        close_btn.clicked.connect(self.close)
        btn_layout.addWidget(close_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # 원본 데이터 에러 강조 테이블
        self.result_label = QLabel("에러 발생 행 (규칙 위반 셀 강조 표시)")
        self.result_label.setStyleSheet("font-weight: bold; color: #c0392b; padding: 4px 0;")
        self.result_label.hide()
        layout.addWidget(self.result_label)

        self.result_table = QTableWidget()
        self.result_table.setColumnCount(0)
        self.result_table.horizontalHeader().setStretchLastSection(True)
        self.result_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.result_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.result_table.setMinimumHeight(300)
        self.result_table.hide()
        layout.addWidget(self.result_table)

        self.setLayout(layout)
    
    def select_csv_file(self):
        """CSV 또는 Parquet 파일 선택"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "데이터 파일 선택", "", 
            "Data Files (*.csv *.parquet);;CSV Files (*.csv);;Parquet Files (*.parquet);;All Files (*.*)"
        )
        if file_path:
            self.csv_file = file_path
            self.file_path_input.setText(file_path)
    
    def load_default_template(self):
        """기본 템플릿 로드"""
        template = """ignit_status,valid_values=[0;1],점화 상태 (0: OFF, 1: ON)
soc_display_rate,range=[0;100],표시 SOC (%)
soh_rate,range=[0;100],SOH (%)
pack_volt,range=[200;500],팩 전압 (V)
pack_curr,range=[-500;500],팩 전류 (A)
module_min_temp,range=[-40;85],최저 모듈 온도 (°C)
module_max_temp,range=[-40;85],최고 모듈 온도 (°C)
cell_max_volt,range=[2.5;4.5],최대 셀 전압 (V)
cell_min_volt,range=[2.5;4.5],최소 셀 전압 (V)
main_relay_status,valid_values=[0;1],메인 릴레이 상태"""
        
        self.rules_text.setPlainText(template)
    
    def parse_rules(self):
        """텍스트 규칙을 파싱"""
        rules_text = self.rules_text.toPlainText().strip()
        if not rules_text:
            return {}
        
        rules = {}
        lines = rules_text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            
            parts = line.split(',')
            if len(parts) < 2:
                continue
            
            column = parts[0].strip()
            check_type = parts[1].strip()
            description = parts[2].strip() if len(parts) > 2 else ''
            
            rule = {
                'check_null': True,
                'description': description
            }
            
            # valid_values 파싱
            if 'valid_values=' in check_type:
                values_str = check_type.split('=')[1].strip('[]')
                values = [v.strip() for v in values_str.split(';')]
                # 숫자로 변환 시도
                try:
                    values = [int(v) if v.isdigit() else float(v) if '.' in v else v for v in values]
                except:
                    pass
                rule['valid_values'] = values
            
            # range 파싱
            elif 'range=' in check_type:
                range_str = check_type.split('=')[1].strip('[]')
                min_val, max_val = range_str.split(';')
                rule['check_range'] = True
                rule['min_value'] = float(min_val.strip())
                rule['max_value'] = float(max_val.strip())
            
            # null만 체크
            elif check_type == 'null':
                rule['check_null'] = True
            
            rules[column] = rule
        
        return rules
    
    def run_validation(self):
        """검증 실행"""
        if not self.csv_file:
            QMessageBox.warning(self, "오류", "데이터 파일을 선택하세요.")
            return
        
        # 파일 형식 확인
        file_ext = os.path.splitext(self.csv_file)[1].lower()
        if file_ext not in ['.csv', '.parquet']:
            QMessageBox.warning(self, "오류", "CSV 또는 Parquet 파일만 지원합니다.")
            return
        
        # 규칙 파싱
        rules = self.parse_rules()
        
        if not rules:
            QMessageBox.warning(self, "오류", "검증 규칙을 입력하세요.")
            return
        
        # 검증 실행
        try:
            validator = BMSDataValidator()
            validator.AWS_ACCESS_KEY_ID = AWS_ACCESS_KEY_ID  # type: ignore
            validator.AWS_SECRET_ACCESS_KEY = AWS_SECRET_ACCESS_KEY  # type: ignore
            
            # validate_file 메서드 사용 (CSV, Parquet 모두 지원)
            results = validator.validate_file(self.csv_file)
            
            # 결과를 DataFrame으로 변환
            result_df = pd.DataFrame(results)
            
            # 결과 저장
            base_name = os.path.splitext(os.path.basename(self.csv_file))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"validation_result_{base_name}_{timestamp}.csv"
            
            result_df.to_csv(output_path, index=False, encoding='utf-8-sig')
            
            # 결과 표시
            file_type = "CSV" if file_ext == '.csv' else "Parquet"
            msg = f"검증 완료! ({file_type} 파일)\n\n"
            msg += f"검사 항목: {len(result_df)}개\n"
            if 'status' in result_df.columns:
                msg += f"PASS: {(result_df['status'] == 'PASS').sum()}개\n"
                msg += f"FAIL: {(result_df['status'] == 'FAIL').sum()}개\n"
                msg += f"WARNING: {(result_df['status'] == 'WARNING').sum()}개\n"
            msg += f"\n결과 파일: {output_path}"
            
            QMessageBox.information(self, "검증 완료", msg)

            # 원본 데이터 로드 → 에러 셀 탐지 → 강조 표시
            try:
                raw_df = pd.read_parquet(self.csv_file) if file_ext == '.parquet' else pd.read_csv(self.csv_file)
                error_mask = self._compute_error_mask(raw_df, rules)
                self._show_raw_data_with_errors(raw_df, error_mask, rules)
                xlsx_path = self._export_highlighted_xlsx(raw_df, error_mask, base_name, timestamp)
                if xlsx_path:
                    msg2 = f"에러 셀 강조 XLSX 저장 완료:\n{xlsx_path}"
                    QMessageBox.information(self, "XLSX 내보내기", msg2)
            except Exception as e2:
                QMessageBox.warning(self, "에러 강조 실패", f"에러 강조 처리 중 문제 발생:\n{str(e2)}")

            # 결과 파일 열기 옵션
            reply = QMessageBox.question(
                self, "결과 확인",
                "결과 파일을 여시겠습니까?",
                QMessageBox.Yes | QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                os.startfile(output_path)

        except Exception as e:
            QMessageBox.critical(self, "오류", f"검증 실패:\n{str(e)}")

    def _compute_error_mask(self, df, rules):
        """규칙에 따라 에러 셀 위치를 Boolean DataFrame으로 반환"""
        error_mask = pd.DataFrame(False, index=df.index, columns=df.columns)
        for col_name, rule in rules.items():
            if col_name not in df.columns:
                continue
            col = df[col_name]
            mask = pd.Series(False, index=df.index)

            if rule.get('check_null'):
                mask = mask | col.isna()

            if 'valid_values' in rule:
                try:
                    valid_num = set(float(v) for v in rule['valid_values'])
                    numeric = pd.to_numeric(col, errors='coerce')
                    mask = mask | (col.notna() & ~numeric.isin(valid_num))
                except Exception:
                    mask = mask | (col.notna() & ~col.isin(rule['valid_values']))

            if rule.get('check_range'):
                try:
                    numeric = pd.to_numeric(col, errors='coerce')
                    out = col.notna() & ((numeric < rule['min_value']) | (numeric > rule['max_value']))
                    mask = mask | out
                except Exception:
                    pass

            error_mask[col_name] = mask
        return error_mask

    def _show_raw_data_with_errors(self, df, error_mask, rules):
        """원본 데이터 중 에러 행만 QTableWidget에 표시 (에러 셀 빨간 배경)"""
        from PyQt5.QtWidgets import QTableWidgetItem
        from PyQt5.QtGui import QColor

        ERROR_BG = QColor('#ffcccc')
        ERROR_FG = QColor('#8b0000')
        NORMAL_BG = QColor('#ffffff')
        NORMAL_FG = QColor('#222222')

        error_row_mask = error_mask.any(axis=1)
        df_err = df[error_row_mask].copy()
        em_err = error_mask[error_row_mask].copy()

        if df_err.empty:
            self.result_label.hide()
            self.result_table.hide()
            return

        # 표시할 컬럼: signal_kst_ts + 규칙 컬럼만 (가독성)
        rule_cols = [c for c in rules.keys() if c in df.columns]
        context_cols = [c for c in ['signal_kst_ts', 'unix_time'] if c in df.columns and c not in rule_cols]
        display_cols = context_cols + rule_cols

        df_show = df_err[display_cols].reset_index(drop=False)
        em_show = em_err[display_cols].reset_index(drop=True)
        all_cols = ['index'] + display_cols
        MAX_ROWS = 2000

        self.result_table.setColumnCount(len(all_cols))
        self.result_table.setHorizontalHeaderLabels(all_cols)
        row_count = min(len(df_show), MAX_ROWS)
        self.result_table.setRowCount(row_count)

        for row_idx in range(row_count):
            for col_idx, col_name in enumerate(all_cols):
                if col_name == 'index':
                    val = str(df_show['index'].iloc[row_idx])
                    item = QTableWidgetItem(val)
                    item.setBackground(NORMAL_BG)
                    item.setForeground(NORMAL_FG)
                else:
                    val = str(df_show[col_name].iloc[row_idx])
                    item = QTableWidgetItem(val)
                    is_error = bool(em_show[col_name].iloc[row_idx]) if col_name in em_show.columns else False
                    if is_error:
                        item.setBackground(ERROR_BG)
                        item.setForeground(ERROR_FG)
                    else:
                        item.setBackground(NORMAL_BG)
                        item.setForeground(NORMAL_FG)
                self.result_table.setItem(row_idx, col_idx, item)

        self.result_table.resizeColumnsToContents()
        total_err = int(error_row_mask.sum())
        self.result_label.setText(
            f"에러 발생 행: {total_err:,}행 (표시: {row_count}행) · 규칙 위반 셀 빨간색 강조"
        )
        self.result_label.show()
        self.result_table.show()
        if self.height() < 950:
            self.resize(self.width(), 950)

    def _export_highlighted_xlsx(self, df, error_mask, base_name, timestamp):
        """에러 셀을 빨간 배경으로 강조한 XLSX 파일 생성"""
        try:
            MAX_ROWS = 100000
            df_out = df.iloc[:MAX_ROWS].reset_index(drop=True)
            em_out = error_mask.iloc[:MAX_ROWS].reset_index(drop=True)

            def highlight(data):
                styles = pd.DataFrame('', index=data.index, columns=data.columns)
                for col in em_out.columns:
                    if col in styles.columns:
                        styles.loc[em_out[col].values, col] = 'background-color: #ffcccc; color: #8b0000'
                return styles

            xlsx_path = f"validation_highlighted_{base_name}_{timestamp}.xlsx"
            df_out.style.apply(highlight, axis=None).to_excel(xlsx_path, engine='openpyxl', index=False)
            return xlsx_path
        except Exception:
            return None


class WorkerThread(QThread):
    """백그라운드 작업을 위한 워커 쓰레드"""
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)
    finished_signal = pyqtSignal(bool, str)
    issues_signal = pyqtSignal(list)

    def __init__(self, task_type, core, params):
        super().__init__()
        self.task_type = task_type
        self.core = core
        self.params = params

    def run(self):
        try:
            if self.task_type == "download":
                self._download_s3()
            elif self.task_type == "download_batch":
                self._download_batch()
            elif self.task_type == "convert":
                self._convert()
            elif self.task_type == "validate":
                self._validate()
            elif self.task_type == "save_report":
                self._save_report()
            elif self.task_type == "merge_files":
                self._merge_files()
        except Exception as e:
            self.log_signal.emit(f"[ERROR] {str(e)}")
            import traceback
            self.log_signal.emit(f"[ERROR] {traceback.format_exc()}")
            self.finished_signal.emit(False, str(e))

    def _download_s3(self):
        import boto3
        
        download_folder = self.params.get('download_folder', False)
        
        if download_folder:
            # 폴더 다운로드
            self._download_s3_folder()
        else:
            # 단일 파일 다운로드
            self._download_s3_file()
    
    def _download_s3_file(self):
        """단일 파일 다운로드"""
        self.log_signal.emit("[INFO] S3 파일 다운로드 시작...")
        self.progress_signal.emit(20)
        
        # S3 핸들러 초기화 (하드코딩된 AWS 키 사용)
        self.core.init_s3(
            bucket=self.params.get('bucket'),
            access_key=AWS_ACCESS_KEY_ID,
            secret_key=AWS_SECRET_ACCESS_KEY,
            region=AWS_DEFAULT_REGION
        )
        
        self.progress_signal.emit(40)
        
        # 다운로드 실행
        self.core.download_from_s3(
            s3_key=self.params['s3_key'],
            local_path=self.params['local_path']
        )
        
        self.progress_signal.emit(100)
        self.log_signal.emit("[SUCCESS] S3 파일 다운로드 완료")
        self.finished_signal.emit(True, "다운로드 완료")
    
    def _download_s3_folder(self):
        """폴더 전체 다운로드"""
        import boto3
        
        self.log_signal.emit("[INFO] S3 폴더 다운로드 시작...")
        self.progress_signal.emit(5)
        
        bucket = self.params.get('bucket')
        s3_prefix = self.params.get('s3_prefix')
        
        # boto3 클라이언트 생성
        s3_client = boto3.client(
            's3',
            aws_access_key_id=AWS_ACCESS_KEY_ID,
            aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
            region_name=AWS_DEFAULT_REGION
        )
        
        try:
            # S3 prefix로 파일 목록 조회
            self.log_signal.emit(f"[INFO] 폴더 내 파일 목록 조회 중...")
            
            paginator = s3_client.get_paginator('list_objects_v2')
            pages = paginator.paginate(Bucket=bucket, Prefix=s3_prefix)
            
            files_to_download = []
            for page in pages:
                if 'Contents' in page:
                    for obj in page['Contents']:
                        key = obj['Key']
                        # 폴더 자체는 제외 (끝에 /로 끝나는 것)
                        if not key.endswith('/'):
                            files_to_download.append(key)
            
            if not files_to_download:
                self.log_signal.emit("[WARNING] 다운로드할 파일이 없습니다.")
                self.finished_signal.emit(False, "파일이 없습니다")
                return
            
            self.log_signal.emit(f"[INFO] 총 {len(files_to_download)}개 파일 발견")
            self.progress_signal.emit(10)
            
            # 다운로드 디렉토리 생성
            download_dir = "downloads"
            os.makedirs(download_dir, exist_ok=True)
            
            success_count = 0
            failed_files = []
            
            for i, s3_key in enumerate(files_to_download):
                try:
                    # 파일명 추출 (prefix 제거)
                    relative_path = s3_key[len(s3_prefix):].lstrip('/')
                    file_name = os.path.basename(s3_key)
                    
                    # 로컬 경로 구성
                    local_path = os.path.join(download_dir, file_name)
                    
                    # 다운로드
                    s3_client.download_file(bucket, s3_key, local_path)
                    self.log_signal.emit(f"[SUCCESS] ({i+1}/{len(files_to_download)}) {file_name}")
                    success_count += 1
                    
                except Exception as e:
                    self.log_signal.emit(f"[ERROR] 실패: {file_name} - {str(e)}")
                    failed_files.append(file_name)
                
                # 진행률 업데이트
                progress = 10 + int(85 * (i + 1) / len(files_to_download))
                self.progress_signal.emit(progress)
            
            self.progress_signal.emit(100)
            
            # 결과 요약
            self.log_signal.emit(f"[INFO] =====================================")
            self.log_signal.emit(f"[INFO] 다운로드 완료: {success_count}/{len(files_to_download)}개")
            if failed_files:
                self.log_signal.emit(f"[WARNING] 실패한 파일: {len(failed_files)}개")
                for fname in failed_files[:5]:  # 최대 5개만 표시
                    self.log_signal.emit(f"[WARNING]   - {fname}")
            
            self.finished_signal.emit(True, f"{success_count}개 파일 다운로드 완료")
            
        except Exception as e:
            self.log_signal.emit(f"[ERROR] 폴더 다운로드 실패: {str(e)}")
            raise

    def _convert(self):
        self.log_signal.emit("[INFO] Parquet → CSV 변환 시작...")
        self.progress_signal.emit(30)
        
        self.core.convert_parquet_to_csv(
            parquet_path=self.params['parquet_path'],
            csv_path=self.params['csv_path']
        )
        
        self.progress_signal.emit(100)
        self.log_signal.emit("[SUCCESS] CSV 변환 완료")
        self.finished_signal.emit(True, "변환 완료")

    def _validate(self):
        self.log_signal.emit("[INFO] 데이터 검증 시작...")
        self.progress_signal.emit(50)
        
        issues = self.core.validate_data()
        
        self.progress_signal.emit(100)
        
        if issues:
            self.log_signal.emit(f"[WARNING] 검증 이슈 발견: {len(issues)}건")
            self.issues_signal.emit(issues)
        else:
            self.log_signal.emit("[SUCCESS] 검증 이슈 없음")
            self.issues_signal.emit([])
        
        self.finished_signal.emit(True, "검증 완료")

    def _save_report(self):
        self.log_signal.emit("[INFO] 리포트 저장 시작...")
        self.progress_signal.emit(50)
        
        self.core.save_report(report_path=self.params['report_path'])
        
        self.progress_signal.emit(100)
        self.log_signal.emit(f"[SUCCESS] 리포트 저장 완료: {self.params['report_path']}")
        self.finished_signal.emit(True, "리포트 저장 완료")
    
    def _merge_files(self):
        """파일 병합 처리"""
        import pandas as pd
        
        self.log_signal.emit("[INFO] 파일 병합 시작...")
        self.progress_signal.emit(10)
        
        file_list = self.params['file_list']
        output_path = self.params['output_path']
        file_type = self.params['file_type']
        
        try:
            df_list = []
            
            if file_type == 'csv':
                # CSV 파일 병합
                self.log_signal.emit(f"[INFO] {len(file_list)}개의 CSV 파일 병합 중...")
                for i, file_path in enumerate(file_list):
                    self.log_signal.emit(f"[INFO] 읽는 중: {os.path.basename(file_path)}")
                    df = pd.read_csv(file_path)
                    df_list.append(df)
                    self.progress_signal.emit(10 + int(40 * (i + 1) / len(file_list)))
                
                # 병합
                merged_df = pd.concat(df_list, ignore_index=True)
                self.progress_signal.emit(60)
                
                # CSV로 저장
                merged_df.to_csv(output_path, index=False, encoding='utf-8-sig')
                self.progress_signal.emit(100)
                self.log_signal.emit(f"[SUCCESS] CSV 병합 완료: {output_path}")
                self.log_signal.emit(f"[INFO] 총 {len(merged_df)}개 행 생성")
                
            elif file_type == 'parquet':
                # Parquet 파일 병합 후 CSV로 저장
                self.log_signal.emit(f"[INFO] {len(file_list)}개의 Parquet 파일 병합 중...")
                
                for i, file_path in enumerate(file_list):
                    self.log_signal.emit(f"[INFO] 읽는 중: {os.path.basename(file_path)}")
                    df = pd.read_parquet(file_path)
                    
                    # 파일명에서 날짜 추출 시도 (정렬용)
                    file_name = os.path.basename(file_path)
                    df['_source_file'] = file_name
                    
                    df_list.append(df)
                    self.progress_signal.emit(10 + int(40 * (i + 1) / len(file_list)))
                
                # 병합
                merged_df = pd.concat(df_list, ignore_index=True)
                self.progress_signal.emit(60)
                
                # 날짜 컬럼이 있으면 정렬
                date_columns = ['timestamp', 'date', 'datetime', 'time', 'created_at']
                sort_column = None
                for col in date_columns:
                    if col in merged_df.columns:
                        sort_column = col
                        break
                
                if sort_column:
                    self.log_signal.emit(f"[INFO] '{sort_column}' 기준으로 정렬 중...")
                    merged_df = merged_df.sort_values(by=sort_column)
                    self.progress_signal.emit(80)
                
                # _source_file 컬럼 제거
                if '_source_file' in merged_df.columns:
                    merged_df = merged_df.drop(columns=['_source_file'])
                
                # CSV로 저장
                merged_df.to_csv(output_path, index=False, encoding='utf-8-sig')
                self.progress_signal.emit(100)
                self.log_signal.emit(f"[SUCCESS] Parquet 병합 완료 (CSV로 저장): {output_path}")
                self.log_signal.emit(f"[INFO] 총 {len(merged_df)}개 행 생성")
            
            self.finished_signal.emit(True, "병합 완료")
            
        except Exception as e:
            self.log_signal.emit(f"[ERROR] 병합 실패: {str(e)}")
            raise
    
    def _download_batch(self):
        """PID와 날짜 기반 배치 다운로드"""
        import boto3
        from datetime import datetime, timedelta
        
        self.log_signal.emit("[INFO] 배치 다운로드 시작...")
        self.progress_signal.emit(5)
        
        env = self.params['env']
        server_type = self.params['server_type']
        pid_list = self.params['pid_list']
        start_date = self.params['start_date']
        end_date = self.params['end_date']
        
        # boto3 클라이언트 생성
        s3_client = boto3.client(
            's3',
            aws_access_key_id=AWS_ACCESS_KEY_ID,
            aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
            region_name=AWS_DEFAULT_REGION
        )
        
        # 버킷명 결정: BCP는 별도 버킷, D2는 DVAL 버킷
        if server_type.lower() == 'bcp':
            bucket_name = f'bcp-{env}-d2-storage-private-apne2'
            self.log_signal.emit(f"[INFO] BCP 버킷 사용: {bucket_name}")
        else:
            bucket_name = 'eplat-validation-monitor'
            self.log_signal.emit(f"[INFO] D2 버킷 사용: {bucket_name}")
        
        # 날짜 범위 생성
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")
        date_list = []
        current = start
        while current <= end:
            date_list.append(current.strftime("%Y-%m-%d"))
            current += timedelta(days=1)
        
        total_files = len(pid_list) * len(date_list)
        downloaded = 0
        success_count = 0
        
        self.log_signal.emit(f"[INFO] 다운로드 대상: PID {len(pid_list)}개 x 날짜 {len(date_list)}일 = {total_files}개 파일")
        
        # 다운로드 실행 날짜 (오늘)
        today = datetime.now().strftime("%Y-%m-%d")
        
        # downloads 디렉토리 생성 (run.bat 위치 기준)
        download_base = os.path.join(os.getcwd(), "downloads", today)
        os.makedirs(download_base, exist_ok=True)
        
        self.log_signal.emit(f"[INFO] 저장 위치: {download_base}")
        
        # PID별 다운로드 성공 여부 추적
        pid_download_status = {pid: {'success': 0, 'files': []} for pid in pid_list}
        
        # 각 PID와 날짜 조합으로 다운로드
        for pid in pid_list:
            # PID 폴더 생성: downloads/오늘날짜/PID/
            pid_folder = os.path.join(download_base, str(pid))
            os.makedirs(pid_folder, exist_ok=True)
            
            for date_str in date_list:
                try:
                    # S3 경로 구성 - BCP와 D2는 경로 구조가 다름
                    if server_type.lower() == 'bcp':
                        # BCP: obd_co_id=MACRIOT/pid=2112/signal_kst_date=2026-01-09/
                        # MACRIOT으로 먼저 검색, 없으면 LGES로 폴백
                        s3_prefix = f"obd_co_id=MACRIOT/pid={pid}/signal_kst_date={date_str}/"
                        self.log_signal.emit(f"[INFO] BCP 경로 검색 (MACRIOT): {s3_prefix}")
                    else:
                        # D2: DVAL/stag 또는 prod/{env}/pid_{PID}/YYYY-MM-DD/
                        s3_prefix = f"DVAL/{server_type}/{env}/pid_{pid}/{date_str}/"
                        self.log_signal.emit(f"[INFO] D2 경로 검색: {s3_prefix}")
                    
                    # 로컬 폴더: downloads/오늘날짜/PID/ (날짜 폴더 없음)
                    local_folder = pid_folder
                    
                    # S3에서 해당 경로의 파일 목록 조회
                    response = s3_client.list_objects_v2(Bucket=bucket_name, Prefix=s3_prefix)
                    
                    # BCP에서 MACRIOT 경로에 파일 없으면 LGES로 폴백
                    if server_type.lower() == 'bcp' and 'Contents' not in response:
                        s3_prefix_lges = f"obd_co_id=LGES/pid={pid}/signal_kst_date={date_str}/"
                        self.log_signal.emit(f"[INFO] MACRIOT 경로 없음 → LGES 폴백: {s3_prefix_lges}")
                        response = s3_client.list_objects_v2(Bucket=bucket_name, Prefix=s3_prefix_lges)
                        s3_prefix = s3_prefix_lges
                    
                    if 'Contents' in response:
                        file_count = 0
                        for obj in response['Contents']:
                            s3_key = obj['Key']
                            file_name = os.path.basename(s3_key)
                            
                            # 빈 폴더는 스킵
                            if not file_name:
                                continue
                            
                            # 로컬 경로 구성
                            local_path = os.path.join(local_folder, file_name)
                            
                            # 다운로드
                            s3_client.download_file(bucket_name, s3_key, local_path)
                            self.log_signal.emit(f"[SUCCESS] PID {pid}/{date_str}: {file_name}")
                            pid_download_status[pid]['success'] += 1
                            pid_download_status[pid]['files'].append(file_name)
                            success_count += 1
                            file_count += 1
                        
                        if file_count == 0:
                            self.log_signal.emit(f"[WARNING] 파일 없음: PID {pid}, 날짜 {date_str}")
                    else:
                        self.log_signal.emit(f"[WARNING] 경로 없음 (MACRIOT/LGES 모두): PID {pid}, 날짜 {date_str}")
                    
                except Exception as e:
                    self.log_signal.emit(f"[ERROR] PID {pid}, 날짜 {date_str}: {str(e)}")
                    import traceback
                    self.log_signal.emit(f"[ERROR] {traceback.format_exc()}")
                
                downloaded += 1
                progress = 5 + int(90 * downloaded / total_files)
                self.progress_signal.emit(progress)
        
        self.log_signal.emit(f"[SUCCESS] 배치 다운로드 완료: {success_count}개 파일")
        self.log_signal.emit(f"[INFO] 저장 위치: {download_base}")

        # 다운로드 완료 후 자동으로 BMS 검증 수행
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if success_count > 0:
            self.log_signal.emit("")
            self.log_signal.emit("=" * 70)
            self.log_signal.emit("[INFO] BMS 데이터 검증 시작...")
            self.log_signal.emit("=" * 70)
            self._auto_bms_validation(download_base, pid_list, pid_download_status, start_date, end_date)
            self._run_trend_analysis(download_base, pid_list, server_type, ts)

        self.progress_signal.emit(100)
        self.finished_signal.emit(True, f"{success_count}개 파일 다운로드 완료")
    
    def _auto_bms_validation(self, base_folder, pids, pid_download_status, start_date, end_date):
        """
        다운로드된 파일 자동 BMS 검증 및 통합 결과 파일 생성
        :param base_folder: downloads/오늘날짜/
        :param pids: PID 리스트
        :param pid_download_status: PID별 다운로드 정보
        :param start_date: 데이터 시작 날짜
        :param end_date: 데이터 종료 날짜
        """
        from data_validator_bms import BMSDataValidator
        from datetime import datetime
        import pandas as pd
        
        # 통합 결과 저장용
        integrated_results = []
        result_number = 1
        
        total_validated = 0
        
        # validator를 루프 밖에서 한 번만 생성 → S3 vehicle_master.json 중복 호출 방지
        shared_validator = BMSDataValidator()
        shared_validator.AWS_ACCESS_KEY_ID = AWS_ACCESS_KEY_ID  # type: ignore
        shared_validator.AWS_SECRET_ACCESS_KEY = AWS_SECRET_ACCESS_KEY  # type: ignore
        
        for pid in pids:
            folder_path = os.path.join(base_folder, str(pid))
            
            if not os.path.exists(folder_path):
                # 다운로드 실패한 PID
                integrated_results.append({
                    '#': result_number,
                    'PID': pid,
                    'PASS/FAIL': 'FAIL',
                    '비고': '다운로드된 파일 없음'
                })
                result_number += 1
                continue
            
            # 폴더 내 모든 CSV/Parquet 파일 검증
            files = [f for f in os.listdir(folder_path) if f.endswith(('.csv', '.parquet'))]
            
            if len(files) == 0:
                integrated_results.append({
                    '#': result_number,
                    'PID': pid,
                    'PASS/FAIL': 'FAIL',
                    '비고': '검증할 파일 없음'
                })
                result_number += 1
                continue
            
            # PID별 전체 결과 취합
            pid_pass = True
            pid_fail_details = []
            
            for file_name in files:
                file_path = os.path.join(folder_path, file_name)
                
                try:
                    self.log_signal.emit(f"[VALIDATION] 검증 중: PID {pid} - {file_name}")
                    
                    results = shared_validator.validate_file(file_path)
                    
                    # 리포트 저장 경로 (같은 폴더, XLSX + HTML)
                    base_name = os.path.splitext(file_name)[0]
                    report_path = os.path.join(folder_path, f"bms_validation_{base_name}.xlsx")
                    shared_validator.generate_report(report_path, file_path)
                    html_report_path = os.path.join(folder_path, f"bms_validation_{base_name}.html")
                    shared_validator.generate_html_report(html_report_path, file_path)
                    self.log_signal.emit(f"  - HTML 리포트: {os.path.basename(html_report_path)}")
                    
                    # 결과 요약
                    total = len(results)
                    pass_count = sum(1 for r in results if r['Status'] == 'PASS')
                    fail_count = sum(1 for r in results if r['Status'] == 'FAIL')
                    warning_count = sum(1 for r in results if r['Status'] == 'WARNING')
                    
                    self.log_signal.emit(f"[VALIDATION] 검증 완료: {file_name}")
                    self.log_signal.emit(f"  - 총 항목: {total} | ✓ PASS: {pass_count} | ✗ FAIL: {fail_count} | ⚠ WARNING: {warning_count}")
                    self.log_signal.emit(f"  - 리포트: {os.path.basename(report_path)}")
                    
                    # FAIL 항목 수집
                    if fail_count > 0:
                        pid_pass = False
                        fail_items = [r for r in results if r['Status'] == 'FAIL']
                        
                        self.log_signal.emit(f"  - FAIL 항목:")
                        for item in fail_items[:3]:  # 상위 3개만 로그
                            self.log_signal.emit(f"    [{item['ID']}] {item['Column']}: {item['Details']}")
                            pid_fail_details.append(f"[{item['ID']}] {item['Column']}: {item['Details']}")
                        
                        if len(fail_items) > 3:
                            self.log_signal.emit(f"    ... 외 {len(fail_items) - 3}개")
                            # 나머지도 비고에 포함
                            for item in fail_items[3:]:
                                pid_fail_details.append(f"[{item['ID']}] {item['Column']}: {item['Details']}")
                    
                    self.log_signal.emit("")
                    total_validated += 1
                    
                except Exception as e:
                    pid_pass = False
                    error_msg = f"검증 실패: {file_name} - {str(e)}"
                    pid_fail_details.append(error_msg)
                    self.log_signal.emit(f"[ERROR] {error_msg}")
                    import traceback
                    self.log_signal.emit(f"[ERROR] {traceback.format_exc()}")
            
            # PID별 결과 추가
            integrated_results.append({
                '#': result_number,
                'PID': pid,
                'PASS/FAIL': 'PASS' if pid_pass else 'FAIL',
                '비고': '' if pid_pass else '; '.join(pid_fail_details[:5])  # 상위 5개만
            })
            result_number += 1
        
        # 통합 결과 파일 생성
        if len(integrated_results) > 0:
            result_df = pd.DataFrame(integrated_results)
            result_file_path = os.path.join(base_folder, f"통합_검증_결과_{start_date}_to_{end_date}.xlsx")
            html_file_path = os.path.join(base_folder, f"통합_검증_결과_{start_date}_to_{end_date}.html")

            # Excel로 저장 (서식 적용)
            self._save_integrated_result_excel(result_df, result_file_path)
            # HTML로도 저장
            self._save_integrated_result_html(result_df, html_file_path, start_date, end_date)

            self.log_signal.emit("")
            self.log_signal.emit("=" * 70)
            self.log_signal.emit(f"[SUCCESS] 통합 결과 파일 생성: {os.path.basename(result_file_path)}")
            self.log_signal.emit(f"[SUCCESS] 통합 HTML 리포트 생성: {os.path.basename(html_file_path)}")
            self.log_signal.emit(f"[INFO] 위치: {base_folder}")
            self.log_signal.emit("=" * 70)
            
            # 폴더 열기
            import subprocess
            subprocess.Popen(f'explorer "{base_folder}"')
        
        if total_validated > 0:
            self.log_signal.emit(f"[SUCCESS] 총 {total_validated}개 파일 검증 완료")
        else:
            self.log_signal.emit("[WARNING] 검증할 CSV/Parquet 파일을 찾지 못했습니다.")

    def _run_trend_analysis(self, base_folder, pid_list, server_type, ts):
        """다운로드된 parquet 파일들로 트렌드 분석 HTML 생성"""
        try:
            import pandas as pd
            from batch_validator import BatchValidator, DEFAULT_RULES
            from trend_analyzer import TrendAnalyzer

            self.log_signal.emit("")
            self.log_signal.emit("=" * 70)
            self.log_signal.emit("[INFO] 트렌드 분석 시작...")
            self.log_signal.emit("=" * 70)

            bv = BatchValidator(bucket="local", rules=DEFAULT_RULES)
            all_records = []

            for pid in pid_list:
                pid_folder = os.path.join(base_folder, str(pid))
                if not os.path.exists(pid_folder):
                    continue
                parquet_files = sorted(f for f in os.listdir(pid_folder) if f.endswith('.parquet'))
                for fname in parquet_files:
                    fpath = os.path.join(pid_folder, fname)
                    try:
                        df = pd.read_parquet(fpath)
                        # 파일명 파싱: MACRIOT_2793_2026-04-13_v20.parquet
                        name_parts = fname.replace('.parquet', '').split('_')
                        obd_co_id = name_parts[0] if name_parts else ''
                        date_str = name_parts[2] if len(name_parts) > 2 else ''
                        meta = {
                            'pid': str(pid),
                            'vehicle_id': f"{server_type}_{pid}",
                            'server_type': server_type,
                            'obd_co_id': obd_co_id,
                            'date': date_str,
                            's3_key': fname,
                        }
                        records = bv._validate_df(df, meta)
                        all_records.extend(records)
                        self.log_signal.emit(f"[TREND] {fname}: {len(records)}개 규칙 분석")
                    except Exception as e:
                        self.log_signal.emit(f"[WARNING] 트렌드 분석 스킵: {fname} - {e}")

            if not all_records:
                self.log_signal.emit("[WARNING] 트렌드 분석할 데이터가 없습니다.")
                return

            results_df = pd.DataFrame(all_records)
            analyzer = TrendAnalyzer(results_df)
            trend_path = os.path.join(base_folder, f"trend_{ts}.html")
            analyzer.generate_html_report(trend_path)
            self.log_signal.emit(f"[SUCCESS] 트렌드 리포트 생성: {os.path.basename(trend_path)}")
            self.log_signal.emit(f"[INFO] 위치: {trend_path}")

        except Exception as e:
            import traceback
            self.log_signal.emit(f"[ERROR] 트렌드 분석 실패: {e}")
            self.log_signal.emit(f"[ERROR] {traceback.format_exc()}")

    def _save_integrated_result_excel(self, df, output_path):
        """
        통합 결과를 서식이 적용된 Excel 파일로 저장
        :param df: 결과 DataFrame (컬럼: #, PID, PASS/FAIL, 비고)
        :param output_path: 저장 경로
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = "검증 결과"
        
        # DataFrame을 시트에 쓰기
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 테두리 스타일
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. 전체 테두리 적용
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 2. 헤더(1행) 주황색 배경
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # 3. PASS/FAIL 컬럼 색상 적용 (3번째 컬럼)
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        pass_font = Font(color="006100", bold=True)
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fail_font = Font(color="9C0006", bold=True)
        
        for row_idx in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row_idx, column=3)
            if status_cell.value == 'PASS':
                status_cell.fill = pass_fill
                status_cell.font = pass_font
            elif status_cell.value == 'FAIL':
                status_cell.fill = fail_fill
                status_cell.font = fail_font
        
        # 4. 비고 컬럼은 왼쪽 정렬
        for row_idx in range(2, ws.max_row + 1):
            remark_cell = ws.cell(row=row_idx, column=4)
            remark_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 8   # #
        ws.column_dimensions['B'].width = 12  # PID
        ws.column_dimensions['C'].width = 15  # PASS/FAIL
        ws.column_dimensions['D'].width = 80  # 비고
        
        # 행 높이 설정
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 30
        
        # 파일 저장
        wb.save(output_path)

    def _save_integrated_result_html(self, df, output_path, start_date, end_date):
        """
        통합 결과를 브라우저에서 바로 열 수 있는 HTML 파일로 저장
        :param df: 결과 DataFrame (컬럼: #, PID, PASS/FAIL, 비고)
        :param output_path: 저장 경로
        :param start_date: 검증 시작일
        :param end_date: 검증 종료일
        """
        from datetime import datetime as _dt

        total = len(df)
        pass_count = (df['PASS/FAIL'] == 'PASS').sum()
        fail_count = (df['PASS/FAIL'] == 'FAIL').sum()
        pass_pct = round(pass_count / total * 100) if total else 0
        generated_at = _dt.now().strftime('%Y-%m-%d %H:%M:%S')

        # SVG 도넛 계산
        circ = 2 * 3.14159 * 26
        pass_dash = round(circ * pass_pct / 100, 2)
        fail_dash = round(circ * (100 - pass_pct) / 100, 2)

        rows_html = ''
        for _, row in df.iterrows():
            status = row['PASS/FAIL']
            color = '#00c4a0' if status == 'PASS' else '#ff5a5a'
            bg = 'rgba(0,196,160,0.07)' if status == 'PASS' else 'rgba(255,90,90,0.07)'
            remark = str(row.get('비고', '') or '').replace('<', '&lt;').replace('>', '&gt;')
            rows_html += f'''
        <tr style="background:{bg};">
          <td style="text-align:center;color:#8a9abb;">{row["#"]}</td>
          <td style="text-align:center;font-weight:700;">{row["PID"]}</td>
          <td style="text-align:center;">
            <span style="display:inline-block;padding:3px 14px;border-radius:20px;
              background:{color}22;color:{color};font-weight:700;font-size:13px;">{status}</span>
          </td>
          <td style="font-size:12px;color:#8a9abb;max-width:500px;word-break:break-all;">{remark}</td>
        </tr>'''

        html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>DART 통합 검증 결과 {start_date}~{end_date}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:"Segoe UI","Malgun Gothic",sans-serif;background:#0d1b2a;color:#e8f0fe;padding:32px;}}
  h1{{font-size:22px;margin-bottom:4px;}}
  .sub{{font-size:13px;color:#8a9abb;margin-bottom:24px;}}
  .summary{{display:flex;gap:20px;align-items:center;background:#132236;border-radius:14px;padding:20px 28px;margin-bottom:24px;}}
  .donut{{flex-shrink:0;}}
  .stats{{display:flex;gap:16px;flex-wrap:wrap;}}
  .stat{{text-align:center;padding:12px 18px;background:#1a2e45;border-radius:10px;min-width:80px;}}
  .stat .num{{font-size:26px;font-weight:800;}}
  .stat .lbl{{font-size:11px;color:#8a9abb;margin-top:4px;}}
  .pass-col{{color:#00c4a0;}} .fail-col{{color:#ff5a5a;}} .tot-col{{color:#7ab3f5;}}
  .search-bar{{display:flex;gap:10px;margin-bottom:14px;}}
  .search-bar input{{flex:1;background:#132236;border:1px solid rgba(255,255,255,0.1);
    color:#e8f0fe;padding:8px 14px;border-radius:8px;font-size:14px;}}
  .filter-btn{{padding:6px 16px;border-radius:20px;border:1px solid rgba(255,255,255,0.15);
    background:none;color:#e8f0fe;cursor:pointer;font-size:13px;transition:all 0.2s;}}
  .filter-btn.active{{background:#1a6bcc;border-color:#1a6bcc;}}
  table{{width:100%;border-collapse:collapse;background:#132236;border-radius:12px;overflow:hidden;}}
  th{{background:#1a2e45;padding:10px 14px;text-align:left;font-size:12px;color:#8a9abb;letter-spacing:0.5px;}}
  td{{padding:10px 14px;border-bottom:1px solid rgba(255,255,255,0.04);vertical-align:top;}}
  tr:last-child td{{border-bottom:none;}}
  @media print{{body{{background:#fff;color:#000;}}table{{border:1px solid #ccc;}}}}
</style>
</head>
<body>
<h1>DART 통합 검증 결과</h1>
<div class="sub">기간: {start_date} ~ {end_date} &nbsp;|&nbsp; 생성: {generated_at} &nbsp;|&nbsp; 총 {total}건</div>

<div class="summary">
  <div class="donut">
    <svg width="80" height="80" viewBox="0 0 60 60">
      <circle cx="30" cy="30" r="26" fill="none" stroke="#1a2e45" stroke-width="8"/>
      <circle cx="30" cy="30" r="26" fill="none" stroke="#ff5a5a" stroke-width="8"
        stroke-dasharray="{fail_dash} {pass_dash}" stroke-dashoffset="0" transform="rotate(-90 30 30)"/>
      <circle cx="30" cy="30" r="26" fill="none" stroke="#00c4a0" stroke-width="8"
        stroke-dasharray="{pass_dash} {fail_dash}" stroke-dashoffset="{fail_dash}" transform="rotate(-90 30 30)"/>
      <text x="30" y="34" text-anchor="middle" font-size="12" font-weight="800" fill="#e8f0fe">{pass_pct}%</text>
    </svg>
  </div>
  <div class="stats">
    <div class="stat"><div class="num tot-col">{total}</div><div class="lbl">전체</div></div>
    <div class="stat"><div class="num pass-col">{pass_count}</div><div class="lbl">PASS</div></div>
    <div class="stat"><div class="num fail-col">{fail_count}</div><div class="lbl">FAIL</div></div>
  </div>
</div>

<div class="search-bar">
  <input id="search" type="text" placeholder="PID 또는 비고 검색..." oninput="filterTable()">
  <button class="filter-btn active" id="btn-ALL" onclick="setFilter('ALL',this)">전체</button>
  <button class="filter-btn" id="btn-PASS" onclick="setFilter('PASS',this)">PASS</button>
  <button class="filter-btn" id="btn-FAIL" onclick="setFilter('FAIL',this)">FAIL</button>
  <button class="filter-btn" onclick="window.print()">인쇄 / PDF</button>
</div>

<table id="tbl">
  <thead><tr><th>#</th><th>PID</th><th>결과</th><th>비고 (FAIL 항목)</th></tr></thead>
  <tbody id="tbody">
{rows_html}
  </tbody>
</table>

<script>
  let activeFilter = 'ALL';
  function setFilter(f, btn) {{
    activeFilter = f;
    document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    filterTable();
  }}
  function filterTable() {{
    const q = document.getElementById('search').value.toLowerCase();
    document.querySelectorAll('#tbody tr').forEach(tr => {{
      const txt = tr.innerText.toLowerCase();
      const cells = tr.querySelectorAll('td');
      const status = cells[2] ? cells[2].innerText.trim() : '';
      const matchFilter = activeFilter === 'ALL' || status === activeFilter;
      const matchSearch = !q || txt.includes(q);
      tr.style.display = matchFilter && matchSearch ? '' : 'none';
    }});
  }}
</script>
</body>
</html>'''

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)


class PrdmrtAppPyQt(QMainWindow):
    def __init__(self):
        super().__init__()
        self.core = PrdmrtCore()
        self.worker = None
        self.validation_issues = []
        self.start_date = None
        self.end_date = None
        
        self.init_ui()
        self.apply_styles()

    def init_ui(self):
        self.setWindowTitle("PRDMRT 데이터 검증 도구")
        self.setGeometry(100, 100, 1000, 700)
        
        # 메뉴바 생성
        self.create_menu_bar()
        
        # 중앙 위젯 설정
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # 1. 서버 설정 그룹
        server_group = self.create_server_group()
        main_layout.addWidget(server_group)
        
        # 2. 차량 PID 입력
        pid_group = self.create_pid_group()
        main_layout.addWidget(pid_group)
        
        # 3. 기간 설정
        date_group = self.create_date_group()
        main_layout.addWidget(date_group)
        
        # 4. Start 버튼 (가운데 정렬)
        start_layout = QHBoxLayout()
        start_layout.addStretch()
        self.start_button = QPushButton("START")
        self.start_button.setMinimumSize(200, 50)
        self.start_button.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                font-size: 18px;
                font-weight: bold;
                border-radius: 10px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
        """)
        self.start_button.clicked.connect(self.start_process)
        start_layout.addWidget(self.start_button)
        start_layout.addStretch()
        main_layout.addLayout(start_layout)
        
        # 5. 로그창 (상태 및 결과)
        log_group = QGroupBox("진행 상황 및 로그")
        log_layout = QVBoxLayout()
        log_group.setLayout(log_layout)
        
        self.log_text = QPlainTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMinimumHeight(100)
        log_layout.addWidget(self.log_text)
        
        main_layout.addWidget(log_group)
        
        # 6. 프로그레스 바
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("%p%")
        self.progress_bar.setMinimumHeight(30)
        main_layout.addWidget(self.progress_bar)

    def create_menu_bar(self):
        """메뉴바 생성"""
        menubar = self.menuBar()
        
        # File 메뉴
        file_menu = menubar.addMenu("File")
        
        load_action = QAction("Load File", self)
        load_action.triggered.connect(self.load_file)
        file_menu.addAction(load_action)
        
        open_dir_action = QAction("Open Directory", self)
        open_dir_action.triggered.connect(self.open_directory)
        file_menu.addAction(open_dir_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Function 메뉴
        function_menu = menubar.addMenu("Function")
        
        s3_download_action = QAction("S3 Download", self)
        s3_download_action.triggered.connect(self.s3_download_dialog)
        function_menu.addAction(s3_download_action)
        
        function_menu.addSeparator()
        
        csv_convert_action = QAction("CSV 변환", self)
        csv_convert_action.triggered.connect(self.convert_to_csv)
        function_menu.addAction(csv_convert_action)
        
        merge_csv_action = QAction("파일 병합 (CSV)", self)
        merge_csv_action.triggered.connect(self.merge_csv_files)
        function_menu.addAction(merge_csv_action)
        
        merge_parquet_action = QAction("파일 병합 (Parquet)", self)
        merge_parquet_action.triggered.connect(self.merge_parquet_files)
        function_menu.addAction(merge_parquet_action)
        
        function_menu.addSeparator()
        
        bms_validation_action = QAction("Data Validation", self)
        bms_validation_action.triggered.connect(self.open_bms_validation)
        function_menu.addAction(bms_validation_action)

    def create_server_group(self):
        """서버 설정 그룹 생성"""
        group = QGroupBox("서버 설정")
        layout = QHBoxLayout()
        
        # 환경 선택 (dev/stag/prod)
        env_label = QLabel("환경:")
        layout.addWidget(env_label)
        
        self.env_group = QButtonGroup()
        self.radio_dev = QRadioButton("DEV")
        self.radio_stag = QRadioButton("STAG")
        self.radio_prod = QRadioButton("PROD")
        self.radio_stag.setChecked(True)  # 기본값: STAG
        
        self.env_group.addButton(self.radio_dev, 1)
        self.env_group.addButton(self.radio_stag, 2)
        self.env_group.addButton(self.radio_prod, 3)
        
        layout.addWidget(self.radio_dev)
        layout.addWidget(self.radio_stag)
        layout.addWidget(self.radio_prod)
        
        layout.addSpacing(40)
        
        # 타입 선택 (D2/BCP)
        type_label = QLabel("타입:")
        layout.addWidget(type_label)
        
        self.type_group = QButtonGroup()
        self.radio_d2 = QRadioButton("D2")
        self.radio_bcp = QRadioButton("BCP")
        self.radio_bcp.setChecked(True)  # 기본값: BCP
        
        self.type_group.addButton(self.radio_d2, 1)
        self.type_group.addButton(self.radio_bcp, 2)
        
        layout.addWidget(self.radio_d2)
        layout.addWidget(self.radio_bcp)
        
        layout.addStretch()
        
        group.setLayout(layout)
        return group

    def create_pid_group(self):
        """차량 PID 입력 그룹"""
        group = QGroupBox("차량 PID")
        layout = QHBoxLayout()
        
        label = QLabel("PID 입력 (쉼표로 구분):")
        layout.addWidget(label)
        
        self.pid_input = QLineEdit()
        self.pid_input.setPlaceholderText("")  # 예시 제거
        layout.addWidget(self.pid_input)
        
        group.setLayout(layout)
        return group

    def create_date_group(self):
        """기간 설정 그룹"""
        group = QGroupBox("기간 설정")
        layout = QHBoxLayout()
        
        # 시작일
        start_label = QLabel("시작일:")
        layout.addWidget(start_label)
        
        self.start_date_input = QLineEdit()
        self.start_date_input.setReadOnly(True)
        self.start_date_input.setPlaceholderText("yyyy-MM-dd")
        layout.addWidget(self.start_date_input)
        
        # 종료일
        end_label = QLabel("종료일:")
        layout.addWidget(end_label)
        
        self.end_date_input = QLineEdit()
        self.end_date_input.setReadOnly(True)
        self.end_date_input.setPlaceholderText("yyyy-MM-dd")
        layout.addWidget(self.end_date_input)
        
        # 캘린더 버튼
        calendar_button = QPushButton("캘린더 선택")
        calendar_button.clicked.connect(self.open_date_picker)
        layout.addWidget(calendar_button)
        
        group.setLayout(layout)
        return group

    def open_date_picker(self):
        """날짜 선택 다이얼로그 열기"""
        dialog = DateRangeDialog(self)
        if dialog.exec_():
            start_date, end_date = dialog.get_date_range()
            self.start_date_input.setText(start_date)
            self.end_date_input.setText(end_date)
            self.start_date = start_date
            self.end_date = end_date
            self.log(f"[INFO] 기간 설정: {start_date} ~ {end_date}")

    def apply_styles(self):
        """QSS 스타일 적용"""
        style = """
            QMainWindow {
                background-color: #f5f5f5;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #cccccc;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QLineEdit {
                border: 1px solid #cccccc;
                border-radius: 3px;
                padding: 5px;
                background-color: white;
            }
            QLineEdit:focus {
                border: 2px solid #4CAF50;
            }
            QPlainTextEdit {
                border: 1px solid #cccccc;
                border-radius: 3px;
                background-color: white;
                font-family: 'Consolas', monospace;
            }
            QProgressBar {
                border: 2px solid #cccccc;
                border-radius: 5px;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 3px;
            }
            QRadioButton {
                spacing: 5px;
            }
            QRadioButton::indicator {
                width: 15px;
                height: 15px;
            }
        """
        self.setStyleSheet(style)

    def log(self, msg):
        """로그 메시지 출력"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.appendPlainText(f"[{timestamp}] {msg}")
    
    def parse_pids(self, pid_input):
        """
        PID 입력 파싱 (숫자 외 모든 문자로 구분)
        :param pid_input: 입력 문자열
        :return: PID 리스트
        """
        import re
        # 숫자만 추출
        pids = re.findall(r'\d+', pid_input)
        return list(set(pids))  # 중복 제거

    def start_process(self):
        """메인 프로세스 시작"""
        # 입력 검증
        pids = self.pid_input.text().strip()
        if not pids:
            QMessageBox.warning(self, "입력 오류", "차량 PID를 입력하세요.")
            return
        
        if not self.start_date or not self.end_date:
            QMessageBox.warning(self, "입력 오류", "기간을 선택하세요.")
            return
        
        # 환경 및 타입 가져오기
        env = "dev" if self.radio_dev.isChecked() else "stag" if self.radio_stag.isChecked() else "prod"
        server_type = "d2" if self.radio_d2.isChecked() else "bcp"
        
        self.log(f"[INFO] 프로세스 시작")
        self.log(f"[INFO] 환경: {env.upper()}, 타입: {server_type.upper()}")
        self.log(f"[INFO] PID 원본: {pids}")
        
        # PID 리스트로 변환 (숫자 외 문자로 구분)
        pid_list = self.parse_pids(pids)
        self.log(f"[INFO] PID 파싱 결과: {', '.join(pid_list)}")
        self.log(f"[INFO] 기간: {self.start_date} ~ {self.end_date}")
        
        # S3 다운로드 시작
        params = {
            'env': env,
            'server_type': server_type,
            'pid_list': pid_list,
            'start_date': self.start_date,
            'end_date': self.end_date
        }
        
        self.start_worker("download_batch", params)

    def load_file(self):
        """파일 로드"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "파일 선택", "", "All Files (*);;CSV Files (*.csv);;Parquet Files (*.parquet)"
        )
        if file_path:
            self.log(f"[INFO] 파일 로드: {file_path}")
            # TODO: 파일 로드 로직

    def open_directory(self):
        """디렉토리 열기"""
        dir_path = QFileDialog.getExistingDirectory(self, "디렉토리 선택")
        if dir_path:
            self.log(f"[INFO] 디렉토리 열기: {dir_path}")
            os.startfile(dir_path)

    def s3_download_dialog(self):
        """S3 링크 입력받아 다운로드 (파일 또는 폴더)"""
        from PyQt5.QtWidgets import QInputDialog
        
        s3_url, ok = QInputDialog.getText(
            self, "S3 Download", 
            "S3 URL, Key 또는 폴더 경로를 입력하세요:\n"
            "(파일 예: path/file.parquet)\n"
            "(폴더 예: DVAL/bcp/prod/pid_296/2024-01-15/)"
        )
        
        if ok and s3_url:
            s3_url = s3_url.strip()
            
            # 빈 입력 체크
            if not s3_url:
                QMessageBox.warning(self, "입력 오류", "S3 경로를 입력하세요.")
                return
            
            # S3 URL 파싱
            if s3_url.startswith('s3://'):
                # s3://bucket/key 형식
                parts = s3_url.replace('s3://', '').split('/', 1)
                bucket = parts[0]
                s3_key = parts[1] if len(parts) > 1 else ''
            else:
                # key만 입력된 경우 기본 버킷 사용
                bucket = 'eplat-validation-monitor'  # 기본 버킷
                s3_key = s3_url
            
            # 디렉토리 생성
            os.makedirs("downloads", exist_ok=True)
            
            # 폴더인지 파일인지 판단 (끝에 / 있거나 확장자 없으면 폴더로 간주)
            is_folder = s3_key.endswith('/') or '.' not in os.path.basename(s3_key)
            
            if is_folder:
                # 폴더 다운로드
                self.log(f"[INFO] S3 폴더 다운로드 시작")
                self.log(f"[INFO] Bucket: {bucket}")
                self.log(f"[INFO] Prefix: {s3_key}")
                
                params = {
                    'bucket': bucket,
                    's3_prefix': s3_key,
                    'download_folder': True
                }
                
                self.start_worker("download", params)
            else:
                # 단일 파일 다운로드
                file_name = os.path.basename(s3_key)
                local_path = os.path.join("downloads", file_name)
                
                self.log(f"[INFO] S3 파일 다운로드 시작")
                self.log(f"[INFO] Bucket: {bucket}")
                self.log(f"[INFO] Key: {s3_key}")
                self.log(f"[INFO] Local: {local_path}")
                
                params = {
                    'bucket': bucket,
                    's3_key': s3_key,
                    'local_path': local_path,
                    'download_folder': False
                }
                
                self.start_worker("download", params)

    def convert_to_csv(self):
        """CSV 변환"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Parquet 파일 선택", "", "Parquet Files (*.parquet)"
        )
        if file_path:
            csv_path = file_path.replace('.parquet', '.csv')
            params = {
                'parquet_path': file_path,
                'csv_path': csv_path
            }
            self.start_worker("convert", params)

    def merge_csv_files(self):
        """CSV 파일 병합"""
        files, _ = QFileDialog.getOpenFileNames(
            self, "CSV 파일 선택 (병합)", "", "CSV Files (*.csv)"
        )
        if files and len(files) > 1:
            # 저장 위치 선택
            save_path, _ = QFileDialog.getSaveFileName(
                self, "병합된 CSV 저장", "merged.csv", "CSV Files (*.csv)"
            )
            if save_path:
                params = {
                    'file_list': files,
                    'output_path': save_path,
                    'file_type': 'csv'
                }
                self.start_worker("merge_files", params)
        elif files:
            QMessageBox.warning(self, "파일 선택", "2개 이상의 파일을 선택하세요.")

    def merge_parquet_files(self):
        """Parquet 파일 병합 후 CSV로 저장"""
        files, _ = QFileDialog.getOpenFileNames(
            self, "Parquet 파일 선택 (병합)", "", "Parquet Files (*.parquet)"
        )
        if files and len(files) > 1:
            # 저장 위치 선택
            save_path, _ = QFileDialog.getSaveFileName(
                self, "병합된 CSV 저장", "merged.csv", "CSV Files (*.csv)"
            )
            if save_path:
                params = {
                    'file_list': files,
                    'output_path': save_path,
                    'file_type': 'parquet'
                }
                self.start_worker("merge_files", params)
        elif files:
            QMessageBox.warning(self, "파일 선택", "2개 이상의 파일을 선택하세요.")

    def start_worker(self, task_type, params):
        """워커 쓰레드 시작"""
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(self, "경고", "이미 작업이 진행 중입니다.")
            return
        
        self.progress_bar.setValue(0)
        self.worker = WorkerThread(task_type, self.core, params)
        
        # 시그널 연결
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.progress_bar.setValue)
        self.worker.finished_signal.connect(self.on_worker_finished)
        self.worker.issues_signal.connect(self.display_validation_issues)
        
        self.worker.start()

    def on_worker_finished(self, success, message):
        """워커 완료 시 처리"""
        if success:
            QMessageBox.information(self, "완료", message)
        else:
            QMessageBox.critical(self, "오류", f"작업 실패: {message}")
        
        self.progress_bar.setValue(0)

    def display_validation_issues(self, issues):
        """검증 이슈 표시"""
        self.validation_issues = issues
        if issues:
            self.log(f"[WARNING] 검증 이슈:")
            for i, issue in enumerate(issues, 1):
                self.log(f"  {i}. {issue}")
        else:
            self.log("[SUCCESS] 검증 이슈 없음")
    
    def open_data_validation(self):
        """데이터 검증 다이얼로그 열기"""
        dialog = DataValidationDialog(self)
        dialog.exec_()
    
    def open_bms_validation(self):
        """BMS 데이터 검증 (37항목) 다이얼로그 열기"""
        dialog = BMSValidationDialog(self)
        dialog.exec_()


class BMSValidationDialog(QDialog):
    """BMS 데이터 검증 (37항목) 다이얼로그"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("BMS Data Validation (37항목 검증)")
        self.setModal(True)
        self.setMinimumSize(900, 600)
        
        layout = QVBoxLayout()
        
        # 설명
        desc_label = QLabel(
            "<b>BMS 데이터 37개 항목 검증</b><br>"
            "unix_time, ignit_status, chg_conr_status_list, em_speed_kmh, pack_curr, pack_volt 등<br>"
            "37개 항목의 상세 검증 규칙을 자동으로 수행합니다."
        )
        desc_label.setStyleSheet("padding: 10px; background-color: #e3f2fd; border-radius: 5px;")
        layout.addWidget(desc_label)
        
        # 파일 선택
        file_group = QGroupBox("데이터 파일 선택")
        file_layout = QHBoxLayout()
        
        self.file_path_input = QLineEdit()
        self.file_path_input.setReadOnly(True)
        self.file_path_input.setPlaceholderText("CSV 또는 Parquet 파일을 선택하세요...")
        file_layout.addWidget(self.file_path_input)
        
        select_file_btn = QPushButton("파일 선택")
        select_file_btn.clicked.connect(self.select_file)
        file_layout.addWidget(select_file_btn)
        
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)
        
        # 리포트 저장 형식
        format_group = QGroupBox("리포트 저장 형식")
        format_layout = QHBoxLayout()
        
        self.csv_radio = QRadioButton("CSV")
        self.xlsx_radio = QRadioButton("Excel (XLSX) - 서식 적용")
        self.xlsx_radio.setChecked(True)  # 기본값: XLSX
        
        format_layout.addWidget(self.xlsx_radio)
        format_layout.addWidget(self.csv_radio)
        format_layout.addStretch()
        
        format_group.setLayout(format_layout)
        layout.addWidget(format_group)
        
        # 진행 상황
        self.progress_label = QLabel("대기 중...")
        layout.addWidget(self.progress_label)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)
        
        # 결과 표시 영역
        result_group = QGroupBox("검증 결과 요약")
        result_layout = QVBoxLayout()
        
        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        self.result_text.setMaximumHeight(200)
        result_layout.addWidget(self.result_text)
        
        result_group.setLayout(result_layout)
        layout.addWidget(result_group)
        
        # 버튼
        button_layout = QHBoxLayout()
        
        validate_btn = QPushButton("검증 시작")
        validate_btn.clicked.connect(self.start_validation)
        validate_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px; font-weight: bold;")
        button_layout.addWidget(validate_btn)
        
        close_btn = QPushButton("닫기")
        close_btn.clicked.connect(self.close)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def select_file(self):
        """파일 선택"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "BMS 데이터 파일 선택", 
            "", 
            "Data Files (*.csv *.parquet);;CSV Files (*.csv);;Parquet Files (*.parquet)"
        )
        if file_path:
            self.file_path_input.setText(file_path)
    
    def start_validation(self):
        """검증 시작"""
        file_path = self.file_path_input.text()
        
        if not file_path:
            QMessageBox.warning(self, "입력 오류", "파일을 선택해주세요.")
            return
        
        if not os.path.exists(file_path):
            QMessageBox.critical(self, "파일 오류", "선택한 파일이 존재하지 않습니다.")
            return
        
        # 리포트 저장 경로 설정
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_basename = os.path.splitext(os.path.basename(file_path))[0]
        
        if self.xlsx_radio.isChecked():
            output_path = f"bms_validation_{file_basename}_{timestamp}.xlsx"
        else:
            output_path = f"bms_validation_{file_basename}_{timestamp}.csv"
        
        # 진행 상태 업데이트
        self.progress_label.setText("검증 중...")
        self.progress_bar.setValue(20)
        self.result_text.clear()
        QApplication.processEvents()
        
        try:
            # BMS 검증 수행
            from data_validator_bms import BMSDataValidator
            
            validator = BMSDataValidator()
            validator.AWS_ACCESS_KEY_ID = AWS_ACCESS_KEY_ID  # type: ignore
            validator.AWS_SECRET_ACCESS_KEY = AWS_SECRET_ACCESS_KEY  # type: ignore
            
            self.progress_bar.setValue(40)
            self.progress_label.setText("파일 로딩 및 검증 중...")
            QApplication.processEvents()
            
            # 검증 실행
            results = validator.validate_file(file_path)
            
            self.progress_bar.setValue(70)
            self.progress_label.setText("리포트 생성 중...")
            QApplication.processEvents()
            
            # 리포트 저장 (XLSX + HTML)
            validator.generate_report(output_path)
            html_output_path = os.path.splitext(output_path)[0] + '.html'
            validator.generate_html_report(html_output_path, file_path)

            self.progress_bar.setValue(100)
            self.progress_label.setText("완료!")

            # 결과 요약 표시
            total = len(results)
            pass_count = sum(1 for r in results if r['Status'] == 'PASS')
            fail_count = sum(1 for r in results if r['Status'] == 'FAIL')
            warning_count = sum(1 for r in results if r['Status'] == 'WARNING')
            na_count = sum(1 for r in results if r['Status'] == 'N/A')

            summary = f"""
=== 검증 결과 요약 ===

총 검증 항목: {total}개

✅ PASS: {pass_count}개
❌ FAIL: {fail_count}개
⚠️  WARNING: {warning_count}개
➖ N/A: {na_count}개

리포트 저장: {output_path}
HTML 리포트: {html_output_path}

=== 주요 실패 항목 ===
"""
            # FAIL 항목 상위 5개 표시
            fail_items = [r for r in results if r['Status'] == 'FAIL']
            for item in fail_items[:5]:
                summary += f"\n[{item['ID']}] {item['Column']} - {item['Check']}: {item['Details']}"
            
            if len(fail_items) > 5:
                summary += f"\n\n... 외 {len(fail_items) - 5}개 실패 항목"
            
            self.result_text.setPlainText(summary)
            
            QMessageBox.information(
                self, 
                "검증 완료", 
                f"BMS 데이터 검증이 완료되었습니다!\n\n"
                f"PASS: {pass_count}, FAIL: {fail_count}, WARNING: {warning_count}\n\n"
                f"상세 리포트: {output_path}"
            )
            
        except Exception as e:
            self.progress_label.setText("오류 발생")
            self.progress_bar.setValue(0)
            QMessageBox.critical(self, "검증 오류", f"검증 중 오류가 발생했습니다:\n{str(e)}")
            logger.error(f"BMS 검증 오류: {e}", exc_info=True)


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')  # 모던한 스타일 적용
    
    window = PrdmrtAppPyQt()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
